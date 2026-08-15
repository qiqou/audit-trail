"""T8 汇总视图 + 问题台账 + 归档清单用例。

覆盖（对应 TASKS.md T8 验收）：
- summary()：三维汇总数量 = 明细数量（状态/版块/单位）
- export_excel：台账含状态/版本数/证据提示列
- package_project：ZIP 内置归档清单，清单文件数 = 实际文件数，哈希可核对
"""

import zipfile
from pathlib import Path

import pytest
from export import export_excel, package_project
from openpyxl import load_workbook


def _mk_project(proj):
    """造一个多单位/多状态/带附件的项目。"""
    u1 = proj.add_unit("华电XX电厂", "张三")
    u2 = proj.add_unit("华电YY电厂", "张三")
    # 单位1：两条底稿（草稿 + 编制完成）
    i1 = proj.add_issue(u1, "张三", department="营销管理", defect_type="电费回收不及时",
                        defect_desc="A电厂问题1", author="张三", reviewer="李四")
    i2 = proj.add_issue(u1, "张三", department="安全生产", defect_type="安全隐患",
                        defect_desc="A电厂问题2", author="张三")
    proj.change_status(i1, "编制完成", "张三")
    proj.update_issue(i2, "张三", defect_desc="A电厂问题2修改")  # 生成版本
    # 单位2：一条底稿（草稿）
    proj.add_issue(u2, "张三", department="营销管理", defect_type="电费回收不及时",
                   defect_desc="B电厂问题")
    # 给 i1 挂一个附件
    src = proj.root / "证据.txt"
    src.write_text("evidence", encoding="utf-8")
    f = proj.add_file(u1, src, "张三", orig_name="证据.txt")
    proj.link_file(i1, f["id"], "张三")
    return u1, u2, i1, i2


def test_summary_counts_match_details(proj):
    """三维汇总数量与明细一致。"""
    _mk_project(proj)
    s = proj.summary()

    assert s["total"] == 3
    # 状态汇总：编制完成 1（i1）+ 草稿 2（i2、B电厂）
    assert s["by_status"]["编制完成"] == 1
    assert s["by_status"]["草稿"] == 2
    # 版块汇总：营销管理 2 + 安全生产 1
    assert s["by_dept"]["营销管理"] == 2
    assert s["by_dept"]["安全生产"] == 1
    # 单位汇总
    assert s["by_unit"]["华电XX电厂"]["issues"] == 2
    assert s["by_unit"]["华电YY电厂"]["issues"] == 1
    # 单位1 有 1 个附件
    assert s["by_unit"]["华电XX电厂"]["files"] == 1
    # 汇总数 = 各维度加总
    assert sum(s["by_status"].values()) == s["total"]
    assert sum(s["by_dept"].values()) == s["total"]
    assert sum(v["issues"] for v in s["by_unit"].values()) == s["total"]


def test_summary_empty_project(proj):
    """空项目汇总全 0。"""
    s = proj.summary()
    assert s["total"] == 0
    assert s["by_status"] == {}
    assert s["by_dept"] == {}
    assert s["by_unit"] == {}


def test_export_excel_has_status_version_evidence(proj):
    """台账增强：状态/版本数/证据提示列正确。"""
    _mk_project(proj)
    r = export_excel(proj, scope="project", operator="张三")
    wb = load_workbook(r["abs_path"])
    ws = wb.active
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert "状态" in headers
    assert "版本数" in headers
    assert "证据提示" in headers

    # 数据行映射
    header_idx = {h: c for c, h in enumerate(headers, 1)}
    rows = []
    for ri in range(3, ws.max_row + 1):
        rows.append({h: ws.cell(row=ri, column=c).value for h, c in header_idx.items()})
    # 3 条底稿
    assert len(rows) == 3
    # 有附件的底稿（i1）证据提示为空
    with_evidence = [r for r in rows if r["附件数"] > 0]
    assert with_evidence and with_evidence[0]["证据提示"] in (None, "")
    # 无附件底稿标"缺证据"
    no_evidence = [r for r in rows if r["附件数"] == 0]
    assert len(no_evidence) == 2
    assert all(r["证据提示"] == "缺证据" for r in no_evidence)
    # 版本数：i2 修改过一次 → 2 版
    v2 = [r for r in rows if r["缺陷描述"] == "A电厂问题2修改"]
    assert v2 and v2[0]["版本数"] == 2
    # 状态列有值
    assert {r["状态"] for r in rows} == {"草稿", "编制完成"}


def test_package_has_manifest(proj):
    """归档 ZIP 内置归档清单：文件数 = 实际文件数，哈希可核对。"""
    _mk_project(proj)
    r = package_project(proj, scope="all")
    with zipfile.ZipFile(r["abs_path"]) as zf:
        names = zf.namelist()
        # 清单存在
        assert "归档清单.txt" in names
        manifest = zf.read("归档清单.txt").decode("utf-8")
        # 清单里登记的文件数（不含清单自身）
        data_files = [n for n in names if n != "归档清单.txt" and not n.endswith("/")]
        assert f"共 {len(data_files)} 个文件" in manifest
        # 清单包含每个数据文件的路径
        for n in data_files:
            assert n in manifest
        # 清单里每一行含 sha256（64 位 hex）
        lines = [l for l in manifest.splitlines() if "\t" in l]
        assert len(lines) == len(data_files)
        for l in lines:
            _path, _size, sha = l.split("\t")
            assert len(sha) == 64
        # 清单自身可被 zip 正常读取（不是损坏成员）
        assert zf.testzip() is None


def test_package_keeps_same_name_evidence_and_returns_real_collision_name(proj, tmp_path, monkeypatch):
    """同名不同内容的证据都要进入归档；同毫秒生成时返回值必须指向实际 ZIP。"""
    unit_id = proj.add_unit("单位A", "张三")
    issue_id = proj.add_issue(unit_id, "张三", defect_type="问题A")
    first = tmp_path / "first.txt"
    second = tmp_path / "second.txt"
    first.write_text("first", encoding="utf-8")
    second.write_text("second", encoding="utf-8")
    for source in (first, second):
        evidence = proj.add_file(unit_id, source, "张三", orig_name="同名证据.txt")
        proj.link_file(issue_id, evidence["id"], "张三")

    monkeypatch.setattr("export._now_ts", lambda: "20260809_120000_000")
    one = package_project(proj)
    two = package_project(proj)
    assert one["filename"] != two["filename"]
    assert one["filename"] == Path(one["abs_path"]).name
    assert two["filename"] == Path(two["abs_path"]).name
    assert not list((proj.root / "输出").glob("问题汇总_*.xlsx"))

    with zipfile.ZipFile(one["abs_path"]) as archive:
        names = [name for name in archive.namelist() if "同名证据" in name]
        assert len(names) == 2
        assert {archive.read(name) for name in names} == {b"first", b"second"}


def test_package_streams_directly_to_zip_without_staging_copies(proj, tmp_path, monkeypatch):
    """归档不应再复制整套附件或生成 staging 目录；ZIP 内容与清单仍保持可核对。"""
    import export

    unit_id = proj.add_unit("单位A", "张三")
    issue_id = proj.add_issue(unit_id, "张三", defect_type="问题A")
    ordinary = tmp_path / "普通证据.txt"
    ordinary.write_bytes(b"ordinary")
    evidence = proj.add_file(unit_id, ordinary, "张三")
    proj.link_file(issue_id, evidence["id"], "张三")

    folder_source = tmp_path / "证据文件夹"
    folder_source.mkdir()
    (folder_source / "成员.txt").write_bytes(b"folder member")
    folder = proj.add_folder(unit_id, [("成员.txt", folder_source / "成员.txt")], "证据文件夹", "张三")
    proj.link_file(issue_id, folder["id"], "张三")

    monkeypatch.setattr(export.shutil, "copy2", lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("归档不应 staging 复制")))
    monkeypatch.setattr(export.shutil, "copytree", lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("归档不应 staging 复制")))
    monkeypatch.setattr(export.tempfile, "TemporaryDirectory", lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("归档不应创建 staging 目录")))

    result = package_project(proj)
    assert not list((proj.root / "输出").glob("问题汇总_*.xlsx"))
    with zipfile.ZipFile(result["abs_path"]) as archive:
        names = archive.namelist()
        assert any(name.endswith("普通证据.txt") for name in names)
        assert any(name.endswith("证据文件夹/成员.txt") for name in names)
        assert archive.testzip() is None


def test_package_failure_cleans_partial_zip(proj, tmp_path, monkeypatch):
    """直接写 ZIP 失败时不应把未完成归档暴露到输出目录。"""
    import export

    unit_id = proj.add_unit("单位A", "张三")
    issue_id = proj.add_issue(unit_id, "张三", defect_type="问题A")
    source = tmp_path / "证据.txt"
    source.write_bytes(b"evidence")
    evidence = proj.add_file(unit_id, source, "张三")
    proj.link_file(issue_id, evidence["id"], "张三")
    monkeypatch.setattr(export, "_now_ts", lambda: "20260813_120000_000")
    monkeypatch.setattr(export, "_write_streamed_archive_file", lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("模拟磁盘写入失败")))

    with pytest.raises(OSError, match="模拟磁盘写入失败"):
        package_project(proj)
    output_dir = proj.root / "输出"
    assert not list(output_dir.glob("归档_*.zip"))
    assert not list(output_dir.glob(".归档_*.tmp"))


def test_package_rejects_attachment_removed_after_preflight_window(proj, tmp_path):
    """即使预检后附件被外部删除，执行期也必须失败，不得静默生成少证据归档。"""
    unit_id = proj.add_unit("单位A", "张三")
    issue_id = proj.add_issue(unit_id, "张三", defect_type="问题A")
    source = tmp_path / "证据.txt"
    source.write_bytes(b"evidence")
    evidence = proj.add_file(unit_id, source, "张三")
    proj.link_file(issue_id, evidence["id"], "张三")
    (proj.root / evidence["rel_path"]).unlink()

    with pytest.raises(FileNotFoundError, match="附件已缺失"):
        package_project(proj)
    assert not list((proj.root / "输出").glob("归档_*.zip"))


def test_package_rejects_attachment_changed_after_preflight_window(proj, tmp_path):
    """归档执行期重算的摘要必须与预检时登记摘要一致，不能打入被替换证据。"""
    unit_id = proj.add_unit("单位A", "张三")
    issue_id = proj.add_issue(unit_id, "张三", defect_type="问题A")
    source = tmp_path / "证据.txt"
    source.write_bytes(b"original")
    evidence = proj.add_file(unit_id, source, "张三")
    proj.link_file(issue_id, evidence["id"], "张三")
    (proj.root / evidence["rel_path"]).write_bytes(b"replaced")

    with pytest.raises(ValueError, match="附件内容已变化"):
        package_project(proj)
    assert not list((proj.root / "输出").glob("归档_*.zip"))


def test_selected_package_empty_never_falls_back_to_all_units(proj):
    """勾选范围为空是输入错误，不能静默打包全项目。"""
    proj.add_unit("单位A", "张三")
    with pytest.raises(ValueError, match="至少选择一个单位"):
        package_project(proj, scope="selected", unit_ids=[])


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@"])
def test_export_escapes_formula_like_user_text(proj, prefix):
    """公式触发字符开头的用户输入在 Excel 中应保持普通文本。"""
    unit_id = proj.add_unit("单位A", "张三")
    proj.add_issue(unit_id, "张三", defect_type=f"{prefix}HYPERLINK(\"https://invalid\",\"点我\")")
    result = export_excel(proj)
    sheet = load_workbook(result["abs_path"], data_only=False).active
    headers = {sheet.cell(2, col).value: col for col in range(1, sheet.max_column + 1)}
    cell = sheet.cell(3, headers["缺陷定性"])
    assert cell.data_type != "f"
    assert cell.value.startswith(prefix)
