"""底稿编号规则（V3.2）：前缀 + 数字序号 + 后缀，写入数据层全程一致。

- issue_no() 默认纯数字；设置规则后编号 = 前缀 + 序号 + 后缀
- API：GET/POST /api/settings/issue-number 读写 meta，留痕
- 导出 Excel 序号列与归档打包目录名均使用编号（唯一识别码）
"""

import zipfile
from pathlib import Path

import pytest
from export import export_excel, package_project
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from main import app


@pytest.fixture
def client():
    with TestClient(app) as c:
        yield c


def _login(client, name="编号测试员"):
    r = client.post("/api/session", json={"operator": name})
    assert r.status_code == 200
    return r.json()["token"]


def _h(token):
    return {"X-Session": token}


def _open_project(client, t, tmp_path, name="编号项目"):
    raw = tmp_path / name
    r = client.post("/api/project/create", json={"path": str(raw), "name": name}, headers=_h(t))
    assert r.status_code == 200
    return Path(r.json()["path"])


def _add_unit_and_issue(client, t, unit_name="华电XX电厂", seq_hint=""):
    r = client.post("/api/units", json={"name": unit_name}, headers=_h(t))
    uid = r.json()["id"]
    r2 = client.post(f"/api/units/{uid}/issues",
                     json={"defect_type": f"问题{seq_hint}", "defect_desc": f"描述{seq_hint}"}, headers=_h(t))
    return uid, r2.json()["id"]


def test_issue_no_default_and_rule(proj):
    """数据层：默认纯数字；设置前后缀后按规则生成。"""
    assert proj.issue_no(3) == "3"
    proj.set_meta("issue_number_prefix", "A-")
    proj.set_meta("issue_number_suffix", "号")
    assert proj.issue_no(3) == "A-3号"
    assert proj.issue_no(12) == "A-12号"


def test_issue_number_api_read_write(client, tmp_path):
    """API：默认空规则；保存后读回一致并留痕。"""
    t = _login(client)
    _open_project(client, t, tmp_path)

    r = client.get("/api/settings/issue-number", headers=_h(t))
    assert r.status_code == 200
    assert r.json() == {"prefix": "", "suffix": ""}

    r2 = client.post("/api/settings/issue-number",
                     json={"prefix": "底稿-", "suffix": ""}, headers=_h(t))
    assert r2.status_code == 200
    assert r2.json() == {"prefix": "底稿-", "suffix": ""}

    r3 = client.get("/api/settings/issue-number", headers=_h(t))
    assert r3.json() == {"prefix": "底稿-", "suffix": ""}

    logs = client.get("/api/logs", headers=_h(t)).json()
    assert any(log["action"] == "更新编号规则" for log in logs)


def test_export_excel_uses_issue_no(client, tmp_path):
    """导出台账的序号列 = 底稿编号（规则应用后）。"""
    t = _login(client)
    created = _open_project(client, t, tmp_path)
    uid, _iid1 = _add_unit_and_issue(client, t, "华电XX电厂", "1")
    r2 = client.post(f"/api/units/{uid}/issues",
                     json={"defect_type": "问题2", "defect_desc": "描述2"}, headers=_h(t))
    assert r2.status_code == 200
    client.post("/api/settings/issue-number", json={"prefix": "A-", "suffix": "号"}, headers=_h(t))

    from database import AuditProject

    proj = AuditProject(created)
    try:
        r = export_excel(proj, scope="project", operator="编号测试员")
    finally:
        proj.close()
    wb = load_workbook(r["abs_path"])
    ws = wb.active
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: c for c, h in enumerate(headers, 1)}
    seqs = [ws.cell(row=ri, column=idx["序号"]).value for ri in range(3, ws.max_row + 1)]
    assert seqs == ["A-1号", "A-2号"]


def test_package_dir_uses_issue_no(client, tmp_path):
    """归档打包目录名 = 编号 + 定性（不分组时用底稿编号）。"""
    t = _login(client)
    created = _open_project(client, t, tmp_path)
    _add_unit_and_issue(client, t, "华电XX电厂", "1")
    client.post("/api/settings/issue-number", json={"prefix": "底稿", "suffix": ""}, headers=_h(t))

    from database import AuditProject

    proj = AuditProject(created)
    try:
        r = package_project(proj, scope="all")
        with zipfile.ZipFile(r["abs_path"]) as zf:
            names = zf.namelist()
    finally:
        proj.close()
    dirs = [n for n in names if n.endswith("/")]
    assert any("底稿1.问题1" in d for d in dirs), f"目录名应使用编号：{dirs}"
