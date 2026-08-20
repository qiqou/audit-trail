"""审计问题汇总 Excel 的稳定格式写入器。"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SUMMARY_HEADERS = [
    ("seq",             "序号",       6),
    ("unit_name",       "被审计单位", 16),
    ("department",      "所属版块",   14),
    ("category",        "问题分类",   14),
    ("defect_type",     "缺陷定性",   14),
    ("defect_desc",     "缺陷描述",   42),
    ("amount",          "问题金额",   10),
    ("regulation_basis", "制度依据",   30),
    ("suggestion",      "审计建议",   30),
    ("author",          "编写人",     10),
    ("reviewer",        "审核人",     10),
    ("status",          "状态",       8),
    ("version_no",      "版本数",     8),
    ("file_count",      "附件数",     8),
    ("evidence",        "证据提示",   14),
]

_HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),
)
_DATA_FONT = Font(name="微软雅黑", size=10)
_DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
_ALTERNATING_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
_TITLE_FONT = Font(name="微软雅黑", size=12, bold=True)


def _formula_like(value: object) -> bool:
    return isinstance(value, str) and value.startswith(("=", "+", "-", "@"))


def build_summary_workbook(
    project_name: str, rows: list[dict], scope_desc: str, operator: str, exported_at: str,
) -> Workbook:
    """按既有列、样式和公式防护生成工作簿。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "审计问题汇总"
    column_count = len(SUMMARY_HEADERS)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"审计项目：{project_name}    {scope_desc}    导出时间：{exported_at}    导出人：{operator}"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    for column_index, (_key, label, width) in enumerate(SUMMARY_HEADERS, 1):
        cell = sheet.cell(row=2, column=column_index, value=label)
        cell.font, cell.fill, cell.alignment, cell.border = _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGNMENT, _THIN_BORDER
        sheet.column_dimensions[chr(64 + column_index)].width = width

    for row_index, row in enumerate(rows, 3):
        for column_index, (key, _label, _width) in enumerate(SUMMARY_HEADERS, 1):
            value = row.get(key, "")
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if key == "amount" and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
            if _formula_like(value):
                cell.data_type = "s"
            cell.font, cell.alignment, cell.border = _DATA_FONT, _DATA_ALIGNMENT, _THIN_BORDER
            if row_index % 2 == 0:
                cell.fill = _ALTERNATING_FILL

    sheet.freeze_panes = "A3"
    return workbook


def summary_excel_bytes(project_name: str, rows: list[dict], scope_desc: str, operator: str, exported_at: str) -> bytes:
    """生成归档内嵌所需的 Excel 字节，不产生输出目录临时文件。"""
    buffer = io.BytesIO()
    build_summary_workbook(project_name, rows, scope_desc, operator, exported_at).save(buffer)
    return buffer.getvalue()
