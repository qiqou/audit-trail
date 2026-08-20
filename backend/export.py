"""导出 / 打包 / 备份 功能模块。

输出约定：
- 所有导出文件带时间戳后缀（YYYYMMDD_HHMMSS），绝不覆盖旧输出
- 导出文件落在项目 输出/ 目录（随项目走，自包含）
- 备份 .auditbak 落在项目上级目录（备份不应混入项目数据）
"""

import csv
import hashlib
import io
import json
import logging
import os
import re
import shutil
import tempfile
import uuid
import zipfile
from datetime import datetime
from pathlib import Path, PurePosixPath

from config import PROJECT_EXT
from database import ATTACH_DIR, OUT_DIR, SYSTEM_METADATA_NAMES, AuditProject, _now, _safe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from platform_adapter import harden_project

logger = logging.getLogger(__name__)


def _sha256_of_file(path) -> str:
    """计算文件 sha256（归档清单核对用）。"""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _is_system_metadata_member(path: str | Path | PurePosixPath) -> bool:
    """只过滤明确的系统垃圾文件；隐藏名称本身可能是审计证据。"""
    return any(part in SYSTEM_METADATA_NAMES for part in PurePosixPath(path).parts)


def _extracted_attachment_path(extract_root: Path, rel_path: str) -> Path:
    """从不可信备份数据库读取附件路径时，仍限制在已解压的附件库内。"""
    candidate = (extract_root / PurePosixPath(str(rel_path or "").replace("\\", "/"))).resolve()
    attachment_root = (extract_root / ATTACH_DIR).resolve()
    if candidate == attachment_root or not candidate.is_relative_to(attachment_root):
        raise ValueError("备份数据库包含非法附件路径")
    return candidate


def _verify_backup_attachment_snapshot(extract_root: Path, connection) -> None:
    """校验备份数据库登记的每项证据均已解压且内容未被替换。

    老版本可能没有摘要，仍校验路径和存在性以兼容恢复；带摘要的现代项目则
    必须逐项核验。这样恢复成功的项目不会出现“数据库引用的是 A、附件实际
    是 B”或仅在后续健康检查才发现附件缺失的情况。
    """
    import sqlite3

    try:
        rows = connection.execute(
            "SELECT rel_path, mime, sha256, orig_name FROM files ORDER BY id"
        ).fetchall()
    except sqlite3.OperationalError:
        return  # 兼容极早期无附件表的项目，后续打开时会迁移
    for rel_path, mime, expected_sha, orig_name in rows:
        source = _extracted_attachment_path(extract_root, rel_path)
        label = str(orig_name or source.name)
        if mime == "folder":
            if not source.is_dir():
                raise ValueError(f"备份缺少文件夹证据：{label}")
            parts = [f"{relative}\t{_sha256_of_file(member)}" for relative, member in _iter_folder_members(source)]
            actual_sha = hashlib.sha256("\n".join(sorted(parts)).encode("utf-8")).hexdigest()
        else:
            if not source.is_file():
                raise ValueError(f"备份缺少附件：{label}")
            actual_sha = _sha256_of_file(source)
        if expected_sha and actual_sha != expected_sha:
            raise ValueError(f"备份附件摘要不一致：{label}")


def _validate_archive_size(zf: zipfile.ZipFile, label: str) -> list[zipfile.ZipInfo]:
    """校验不可信压缩包的成员数与解压总量，返回已读取的成员列表。"""
    from limits import MAX_ARCHIVE_MEMBERS, MAX_EXTRACT_TOTAL, human_size

    infos = zf.infolist()
    if len(infos) > MAX_ARCHIVE_MEMBERS:
        raise ValueError(f"{label}包含超过 {MAX_ARCHIVE_MEMBERS} 个文件或目录，拒绝处理")
    total = sum(info.file_size for info in infos)
    if total > MAX_EXTRACT_TOTAL:
        raise ValueError(f"{label}解压总量超过上限 {human_size(MAX_EXTRACT_TOTAL)}，拒绝处理")
    return infos

# 汇总表列定义：字段名 → (表头, 列宽)
SUMMARY_HEADERS = [
    ("seq",             "序号",       6),
    ("unit_name",       "被审计单位", 16),
    ("department",      "所属版块",   14),
    ("category",        "问题分类",   14),
    ("defect_type",     "缺陷定性",   14),
    ("defect_desc",     "缺陷描述",   42),
    ("amount",          "问题金额",   10),
    ("regulation_basis","制度依据",   30),
    ("suggestion",      "审计建议",   30),
    ("author",          "编写人",     10),
    ("reviewer",        "审核人",     10),
    ("status",          "状态",       8),
    ("version_no",      "版本数",     8),
    ("file_count",      "附件数",     8),
    ("evidence",        "证据提示",   14),
]

HF = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HFILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HA = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
DF = Font(name="微软雅黑", size=10)
DA = Alignment(vertical="center", wrap_text=True)
AF = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
TITLE_F = Font(name="微软雅黑", size=12, bold=True)


def _now_ts() -> str:
    """时间戳到毫秒，配合 _unique_path 保证输出文件名不重复（审查 F-04 修复）。"""
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]


def _unique_path(out_dir: Path, base_name: str) -> Path:
    """输出路径防覆盖（审查 F-04 修复）：已存在则追加 _1/_2/_3... 序号。

    返回的路径确保当前不存在，绝不覆盖旧输出。
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    p = out_dir / base_name
    if not p.exists():
        return p
    stem = p.stem
    for i in range(1, 10000):
        cand = out_dir / f"{stem}_{i}{p.suffix}"
        if not cand.exists():
            return cand
    raise RuntimeError(f"无法生成唯一文件名：{base_name}")




def _formula_like(value) -> bool:
    """识别可能被电子表格程序解释为公式的用户文本。"""
    return isinstance(value, str) and value.startswith(("=", "+", "-", "@"))


def _ensure_disk_space(target_dir: Path, needed_bytes: int) -> None:
    """解压前磁盘空间检查（I3）：目标盘剩余空间需满足所需 +10% 余量。"""
    target = target_dir if target_dir.is_dir() else target_dir.parent
    usage = shutil.disk_usage(target)
    required = int(needed_bytes * 1.1) + 64 * 1024 * 1024  # 内容 +10% + 64MB 余量
    if usage.free < required:
        raise ValueError(
            f"磁盘剩余空间不足（需约 {required / (1024 ** 3):.1f} GB，"
            f"剩余 {usage.free / (1024 ** 3):.1f} GB），请清理后重试"
        )


def _format_amount_display(amount, currency="", amount_unit=""):
    """金额显示串（审查 I1 修复）：结构化金额拼 "币种 千分位两位小数单位"，避免导出丢失币种/单位。

    - 结构化数字（如 "1234.5" + CNY + 元）→ "CNY 1,234.50元"
    - 无币种/单位的纯数字 → 返回 float，供 Excel 数值单元格 + #,##0.00 格式（可合计）
    - 自由文本（如 "约120万"）→ 原样返回
    """
    raw = str(amount or "").strip()
    if not raw:
        return ""
    try:
        num = float(raw)
    except (TypeError, ValueError):
        return raw
    currency = str(currency or "").strip()
    amount_unit = str(amount_unit or "").strip()
    if not currency and not amount_unit:
        return num
    num_str = f"{num:,.2f}"
    prefix = f"{currency} " if currency else ""
    return f"{prefix}{num_str}{amount_unit}"


def _collect_rows(proj: AuditProject, scope: str = "project", unit_id: int = None,
                   unit_ids: list[int] = None) -> tuple[list[dict], str]:
    """收集汇总行。scope: unit(单单位)/selected(多单位)/project(全部)。返回 (rows, 范围说明)。"""
    rows = []
    if scope == "unit" and unit_id:
        units = [u for u in proj.list_units() if u["id"] == unit_id]
        scope_desc = f"单位：{units[0]['name']}" if units else "单位"
    elif scope == "selected" and unit_ids:
        units = [u for u in proj.list_units() if u["id"] in unit_ids]
        scope_desc = f"勾选单位 {len(units)} 个"
    else:
        units = proj.list_units()
        scope_desc = "全部单位"
    # T8 台账 N+1 优化：一次查询全部版本数，不再每行调 list_versions（审查）
    ver_counts = proj.version_counts()
    issues_by_unit = proj.list_issues_by_unit()
    for u in units:
        for iss in issues_by_unit.get(u["id"], []):
            row = {k: iss.get(k, "") for k, _v, _w in SUMMARY_HEADERS if k not in ("unit_name", "file_count", "version_no", "evidence")}
            # 序号列 = 底稿编号（前缀+序号+后缀），与树/详情一致，作为唯一识别码
            row["seq"] = proj.issue_no(iss["seq"])
            row["unit_name"] = u["name"]
            row["file_count"] = iss.get("file_count", 0)
            # 审查 I1 修复：金额列拼币种/单位（"CNY 1,234.50元"），不再输出裸数字丢币种
            row["amount"] = _format_amount_display(
                iss.get("amount", ""), iss.get("currency", ""), iss.get("amount_unit", ""),
            )
            # 版本数：初始 v1 + 每次修改留版本
            row["version_no"] = ver_counts.get(iss["id"], 1)
            # 证据提示：无附件 → 提示缺证据（T8 台账增强）
            row["evidence"] = "" if row["file_count"] else "缺证据"
            rows.append(row)
    return rows, scope_desc


def _summary_workbook(proj: AuditProject, rows: list[dict], scope_desc: str, operator: str) -> Workbook:
    """按既有汇总样式创建工作簿；文件输出和 ZIP 内嵌共用同一渲染结果。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "审计问题汇总"

    # 标题行
    ncol = len(SUMMARY_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tc = ws.cell(row=1, column=1)
    tc.value = f"审计项目：{proj.project_name}    {scope_desc}    导出时间：{_now()}    导出人：{operator}"
    tc.font = TITLE_F
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # 表头
    for ci, (_k, label, width) in enumerate(SUMMARY_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font, c.fill, c.alignment, c.border = HF, HFILL, HA, THIN
        ws.column_dimensions[chr(64 + ci)].width = width

    # 数据（斑马纹）
    for ri, row in enumerate(rows, 3):
        for ci, (k, _label, _w) in enumerate(SUMMARY_HEADERS, 1):
            v = row.get(k, "")
            c = ws.cell(row=ri, column=ci, value=v)
            if k == "amount" and isinstance(v, (int, float)):
                # 审查 I1 修复：无币种/单位的纯数字金额写数值单元格，可合计且带千分位格式
                c.number_format = "#,##0.00"
            if _formula_like(v):
                # 明确写为字符串，既阻止公式执行，也避免前导单引号污染归档回导后的原始文本。
                c.data_type = "s"
            c.font, c.alignment, c.border = DF, DA, THIN
            if ri % 2 == 0:
                c.fill = AF

    ws.freeze_panes = "A3"
    return wb


def _summary_excel_bytes(proj: AuditProject, rows: list[dict], scope_desc: str, operator: str = "") -> bytes:
    """生成归档内嵌的汇总表，避免先落到“输出”再复制进 ZIP。"""
    buffer = io.BytesIO()
    _summary_workbook(proj, rows, scope_desc, operator).save(buffer)
    return buffer.getvalue()


def export_excel(proj: AuditProject, scope: str = "project", operator: str = "",
                   unit_id: int = None, unit_ids: list[int] = None) -> dict:
    """生成问题汇总表 Excel 到项目 输出/。scope: unit/selected(需 unit_ids) / project。"""
    if scope not in ("unit", "selected", "project"):
        raise ValueError("导出范围无效")
    if scope == "unit" and not unit_id:
        raise ValueError("导出当前单位需指定单位")
    if scope == "selected" and not unit_ids:
        raise ValueError("导出勾选单位需指定单位列表")
    rows, scope_desc = _collect_rows(proj, scope, unit_id, unit_ids)

    out_dir = proj.root / OUT_DIR
    out_dir.mkdir(exist_ok=True)
    suffix = "全部单位" if scope == "project" else ("勾选单位" if scope == "selected" else "当前单位")
    filename = f"问题汇总_{_safe(proj.project_name)}_{suffix}_{_now_ts()}.xlsx"
    # 防覆盖：同秒/同名已存在时自动追加序号（审查 F-04 修复）
    out_path = _unique_path(out_dir, filename)
    _summary_workbook(proj, rows, scope_desc, operator).save(out_path)
    return {"filename": out_path.name, "abs_path": str(out_path), "count": len(rows)}


def export_audit_log_csv(proj: AuditProject, rows: list[dict]) -> dict:
    """导出已筛选的永久操作日志；保持字段稳定且不包含附件或正文内容。"""
    out_dir = proj.root / OUT_DIR
    out_dir.mkdir(exist_ok=True)
    out_path = _unique_path(out_dir, f"操作日志_{_safe(proj.project_name)}_{_now_ts()}.csv")
    fields = ("id", "created_at", "operator", "action", "target", "detail", "event_uuid", "issue_uuid", "file_uuid")
    with out_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in rows)
    return {"filename": out_path.name, "abs_path": str(out_path), "count": len(rows)}


def export_diagnostics_support_package(proj: AuditProject) -> dict:
    """输出可供支持人员读取的最小 JSON 摘要，不包含项目业务数据。"""
    out_dir = proj.root / OUT_DIR
    out_dir.mkdir(exist_ok=True)
    out_path = _unique_path(out_dir, f"审迹诊断支持包_{_now_ts()}.json")
    with out_path.open("w", encoding="utf-8") as handle:
        json.dump(proj.diagnostics_summary(), handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")
    return {"filename": out_path.name}


def _add_archive_dir(zf: zipfile.ZipFile, path: str, reserved: set[str]) -> None:
    """写入空目录，保留无附件问题和空文件夹证据的既有归档结构。"""
    normalized = path.strip("/")
    if normalized and normalized not in reserved:
        zf.writestr(f"{normalized}/", "")
        reserved.add(normalized)


def _unique_archive_path(parent: str, name: str, reserved: set[str]) -> str:
    """给归档内同名附件编号，语义与临时目录中的 ``_unique_path`` 保持一致。"""
    safe_name = _safe(name)
    candidate = f"{parent.rstrip('/')}/{safe_name}"
    if candidate not in reserved:
        reserved.add(candidate)
        return candidate
    suffix = Path(safe_name).suffix
    stem = safe_name[:-len(suffix)] if suffix else safe_name
    for index in range(1, 10_000):
        candidate = f"{parent.rstrip('/')}/{stem}_{index}{suffix}"
        if candidate not in reserved:
            reserved.add(candidate)
            return candidate
    raise RuntimeError(f"无法生成唯一归档路径：{safe_name}")


def _write_streamed_archive_file(zf: zipfile.ZipFile, source: Path, archive_path: str) -> tuple[int, str]:
    """一次读取源文件，同时压缩写入 ZIP 并计算清单摘要。"""
    digest = hashlib.sha256()
    size = 0
    with source.open("rb") as src, zf.open(archive_path, "w", force_zip64=True) as dst:
        while chunk := src.read(1 << 20):
            dst.write(chunk)
            digest.update(chunk)
            size += len(chunk)
    return size, digest.hexdigest()


def _write_folder_to_archive(zf: zipfile.ZipFile, source: Path, archive_root: str,
                             reserved: set[str], manifest_lines: list[str]) -> str:
    """流式写入文件夹证据；沿用既有规则，仅跳过以 . 开头的系统隐藏成员。"""
    # ``_unique_archive_path`` 先占用顶层名字以处理同名文件夹；这里将它转换为
    # 明确目录条目，保留旧实现中空文件夹也可见的 ZIP 结构。
    reserved.discard(archive_root)
    digest_parts: list[str] = []
    for dirpath, dirs, filenames in os.walk(source):
        dirs[:] = [directory for directory in dirs if directory not in SYSTEM_METADATA_NAMES]
        relative_dir = Path(dirpath).relative_to(source)
        archive_dir = archive_root if relative_dir == Path(".") else f"{archive_root}/{relative_dir.as_posix()}"
        _add_archive_dir(zf, archive_dir, reserved)
        for filename in filenames:
            if filename in SYSTEM_METADATA_NAMES:
                continue
            archive_file = _unique_archive_path(archive_dir, filename, reserved)
            size, sha = _write_streamed_archive_file(zf, Path(dirpath) / filename, archive_file)
            manifest_lines.append(f"{archive_file}\t{size}\t{sha}")
            relative = (Path(dirpath) / filename).relative_to(source).as_posix()
            digest_parts.append(f"{relative}\t{sha}")
    return hashlib.sha256("\n".join(sorted(digest_parts)).encode("utf-8")).hexdigest()


def _package_project_direct(proj: AuditProject, scope: str = "all", unit_ids: list[int] = None,
                            group_by_dept: bool = False, operator: str = "") -> dict:
    """直接流式写归档 ZIP，不复制整套附件到 staging 目录。

    目录结构：
      项目名称/                      一级
      附件-单位名称/                 二级
      版块分类/                      （三级，group_by_dept 勾选时启用）
      序号.问题定性/                  末级
    scope: all=全部单位 / selected=勾选单位（unit_ids）
    """
    if scope not in {"all", "selected"}:
        raise ValueError("归档范围无效，请选择全部单位或勾选单位")
    selected_ids = list(dict.fromkeys(unit_ids or []))
    if scope == "selected" and not selected_ids:
        raise ValueError("勾选单位归档时请至少选择一个单位，系统不会自动改为全部单位")
    known_ids = {unit["id"] for unit in proj.list_units()}
    missing_ids = [unit_id for unit_id in selected_ids if unit_id not in known_ids]
    if missing_ids:
        raise ValueError(f"勾选单位不存在或已被删除：{missing_ids}")

    # 单位 → [版块] → 序号.问题定性 → 附件
    units = proj.list_units()
    if scope == "selected":
        units = [u for u in units if u["id"] in selected_ids]
    issues_by_unit = proj.list_issues_by_unit()
    scoped_issue_ids = [issue["id"] for unit in units for issue in issues_by_unit.get(unit["id"], [])]
    files_by_issue = proj.files_for_issues(scoped_issue_ids)
    issue_count = len(scoped_issue_ids)

    out_dir = proj.root / OUT_DIR
    out_dir.mkdir(exist_ok=True)
    filename = f"归档_{_safe(proj.project_name)}_{_now_ts()}.zip"
    # 防覆盖：同秒/同名已存在时自动追加序号（审查 F-04 修复）
    out_path = _unique_path(out_dir, filename)
    temporary_path = out_path.with_name(f".{out_path.name}.{uuid.uuid4().hex}.tmp")
    root = _safe(proj.project_name)
    # 归档清单（T8）：ZIP 内所有文件相对路径 + 大小 + sha256，供接收方核对。
    # 输出 ZIP 先写同目录临时文件，完整关闭后再原子替换，失败不产生假成功归档。
    manifest_lines = [
        "归档清单",
        f"项目：{proj.project_name}",
        f"生成时间：{_now()}",
        f"单位数：{len(units)}    底稿数：{issue_count}",
        "",
        "文件清单（相对路径 / 大小 / sha256）：",
    ]
    reserved: set[str] = set()
    try:
        with zipfile.ZipFile(temporary_path, "w", zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
            _add_archive_dir(zf, root, reserved)
            summary_scope = "selected" if scope == "selected" else "project"
            summary_rows, summary_scope_desc = _collect_rows(proj, summary_scope, unit_ids=selected_ids or None)
            summary_path = _unique_archive_path(root, "审计问题汇总.xlsx", reserved)
            summary_bytes = _summary_excel_bytes(proj, summary_rows, summary_scope_desc, operator)
            zf.writestr(summary_path, summary_bytes, compress_type=zipfile.ZIP_DEFLATED)
            manifest_lines.append(f"{summary_path}\t{len(summary_bytes)}\t{hashlib.sha256(summary_bytes).hexdigest()}")

            for unit in units:
                unit_dir = f"{root}/附件-{_safe(unit['name'])}"
                _add_archive_dir(zf, unit_dir, reserved)
                issues = issues_by_unit.get(unit["id"], [])
                groups = [("", issues)]
                if group_by_dept:
                    grouped: dict[str, list[dict]] = {}
                    for issue in issues:
                        grouped.setdefault(_safe(issue.get("department") or "未分版块"), []).append(issue)
                    groups = list(grouped.items())
                for department, group_issues in groups:
                    parent_dir = unit_dir if not department else f"{unit_dir}/{department}"
                    _add_archive_dir(zf, parent_dir, reserved)
                    for position, issue in enumerate(group_issues, 1):
                        number = position if group_by_dept else issue["seq"]
                        label = f"{proj.issue_no(number)}.{_safe(issue.get('defect_type') or '未定性')}"
                        issue_dir = f"{parent_dir}/{label}"
                        _add_archive_dir(zf, issue_dir, reserved)
                        for evidence in files_by_issue[issue["id"]]:
                            source = proj.attachment_path(evidence["rel_path"])
                            archive_path = _unique_archive_path(issue_dir, evidence["orig_name"], reserved)
                            if evidence.get("mime") == "folder":
                                if not source.is_dir():
                                    raise FileNotFoundError(f"归档时文件夹证据已缺失：{evidence['orig_name']}")
                                actual_sha = _write_folder_to_archive(zf, source, archive_path, reserved, manifest_lines)
                                if evidence.get("sha256") and actual_sha != evidence["sha256"]:
                                    raise ValueError(f"归档时文件夹证据内容已变化：{evidence['orig_name']}")
                            else:
                                if not source.is_file():
                                    raise FileNotFoundError(f"归档时附件已缺失：{evidence['orig_name']}")
                                size, sha = _write_streamed_archive_file(zf, source, archive_path)
                                if evidence.get("sha256") and sha != evidence["sha256"]:
                                    raise ValueError(f"归档时附件内容已变化：{evidence['orig_name']}")
                                manifest_lines.append(f"{archive_path}\t{size}\t{sha}")
            manifest_lines.append(f"\n共 {len(manifest_lines) - 6} 个文件")
            zf.writestr("归档清单.txt", "\n".join(manifest_lines), compress_type=zipfile.ZIP_DEFLATED)
        os.replace(temporary_path, out_path)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise
    return {"filename": out_path.name, "abs_path": str(out_path), "units": len(units),
            "issues": issue_count, "files": len(manifest_lines) - 7}


def package_project(proj: AuditProject, scope: str = "all", unit_ids: list[int] = None,
                    group_by_dept: bool = False, operator: str = "") -> dict:
    """生成完整归档；附件直接流式压缩，避免 staging 目录占用第二份项目空间。"""
    return _package_project_direct(proj, scope=scope, unit_ids=unit_ids,
                                   group_by_dept=group_by_dept, operator=operator)


def archive_preflight(proj: AuditProject, scope: str = "all", unit_ids: list[int] = None,
                      group_by_dept: bool = False) -> dict:
    """归档前执行全量完整性核对，返回阻断项、可确认警告和数据指纹。

    这个函数不写项目数据，确认令牌由 API 会话层保管。打包前会再次调用本函数，
    以防用户完成核对后修改底稿、附件或归档范围。
    """
    if scope not in {"all", "selected"}:
        raise ValueError("归档范围无效，请选择全部单位或勾选单位")
    selected_ids = list(dict.fromkeys(unit_ids or []))
    units = proj.list_units()
    known_ids = {unit["id"] for unit in units}
    if scope == "selected" and not selected_ids:
        raise ValueError("勾选单位归档时请至少选择一个单位，系统不会自动改为全部单位")
    missing_ids = [unit_id for unit_id in selected_ids if unit_id not in known_ids]
    if missing_ids:
        raise ValueError(f"勾选单位不存在或已被删除：{missing_ids}")
    scoped_units = units if scope == "all" else [unit for unit in units if unit["id"] in selected_ids]

    blockers: list[dict] = []
    warnings: list[dict] = []
    if not scoped_units:
        blockers.append({"code": "empty_scope", "message": "归档范围为空，至少需要一个被审计单位"})

    health = proj.health_check(sample_size=0)
    for problem in health["problems"]:
        target = blockers if problem.get("severity") == "P0" else warnings
        target.append({
            "code": str(problem.get("type") or "integrity"),
            "message": str(problem.get("message") or "附件完整性核对失败"),
        })

    unresolved_conflicts = proj.unresolved_merge_conflicts()
    if unresolved_conflicts:
        blockers.append({
            "code": "unresolved_merge_conflicts",
            "message": f"存在 {len(unresolved_conflicts)} 项未确认的合并冲突，须先完成负责人处理后再归档",
        })

    log_chain = proj.verify_audit_log_chain()
    if not log_chain["ok"]:
        blockers.append({
            "code": "audit_log_chain_invalid",
            "message": f"永久操作日志校验失败（{len(log_chain['problems'])} 处），不能归档",
        })

    issues_by_unit = proj.list_issues_by_unit()
    scoped_issues = [issue for unit in scoped_units for issue in issues_by_unit.get(unit["id"], [])]
    status_counts: dict[str, int] = {}
    for issue in scoped_issues:
        status = str(issue.get("status") or "草稿")
        status_counts[status] = status_counts.get(status, 0) + 1
    non_archived = len(scoped_issues) - status_counts.get("已归档", 0)
    if non_archived:
        warnings.append({
            "code": "non_archived_issues",
            "message": f"归档范围内有 {non_archived} 条底稿尚未处于“已归档”状态（{', '.join(f'{name} {count}' for name, count in status_counts.items() if name != '已归档')}）",
        })

    # 指纹覆盖归档范围、底稿当前版本、附件关联及永久日志末端；确认后任一业务
    # 改动都会导致令牌失效。附件物理内容由上面的全量 health_check 覆盖。
    snapshot_rows = {
        "project_uuid": proj.project_uuid,
        "scope": scope,
        "unit_ids": [unit["id"] for unit in scoped_units],
        "group_by_dept": bool(group_by_dept),
        "issues": [
            [issue.get("issue_uuid", ""), issue["id"], issue.get("updated_at", ""), issue.get("status", ""), issue.get("seq", 0)]
            for issue in scoped_issues
        ],
        "files": [
            list(row) for row in proj._conn.execute(
                "SELECT id, file_uuid, rel_path, sha256, size, exclusive_to FROM files "
                "WHERE unit_id IN ({}) ORDER BY id".format(
                    ",".join("?" for _ in scoped_units) or "NULL"
                ),
                tuple(unit["id"] for unit in scoped_units),
            ).fetchall()
        ],
        "links": [
            list(row) for row in proj._conn.execute(
                "SELECT issue_id, file_id FROM issue_files ORDER BY issue_id, file_id"
            ).fetchall()
        ],
        "last_log": list(proj._conn.execute(
            "SELECT id, event_hash FROM audit_log ORDER BY id DESC LIMIT 1"
        ).fetchone() or (0, "")),
    }
    fingerprint = hashlib.sha256(
        json.dumps(snapshot_rows, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    return {
        "ok": not blockers,
        "blockers": blockers,
        "warnings": warnings,
        "counts": {
            "units": len(scoped_units), "issues": len(scoped_issues), "files": health["counts"].get("files", 0),
            "non_archived": non_archived,
        },
        "health": {"checked": health["sample"], "problems": len(health["problems"])},
        "fingerprint": fingerprint,
    }


def create_backup(proj: AuditProject) -> dict:
    """备份：audit.db + 附件库 打包为 .auditbak，存项目上级目录。

    一致性（审查 F-12 修复）：
    - 整个快照和附件写入过程持有项目写锁；工作台的变更请求会等待备份结束
    - 用 sqlite3 backup API 生成一致性数据库快照（不直接复制正在使用的 db）
    - 附件以快照内 ``files`` 清单为准逐项写入并核验 SHA-256，不把未登记的
      临时文件混入备份，也不接受内容变化后仍显示成功的备份
    - ZIP 先写同目录 .tmp 文件，完成后 os.replace 原子落盘（不会出现半成品备份）
    """
    # ``AuditProject`` 的每个业务写入都以同一把锁提交数据库。长备份会暂时
    # 排队写请求，换取可恢复的同一时点。
    with proj._lock:
        return _create_backup_locked(proj)


def _snapshot_backup_files(snapshot_path: Path) -> list[dict]:
    """读取 SQLite 快照中的附件登记，绝不从仍在变化的 live DB 取清单。"""
    import sqlite3

    connection = sqlite3.connect(f"file:{snapshot_path}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    try:
        try:
            rows = connection.execute(
                "SELECT rel_path, mime, sha256, orig_name FROM files ORDER BY id"
            ).fetchall()
        except sqlite3.OperationalError as exc:
            raise ValueError("备份数据库缺少附件登记表") from exc
    finally:
        connection.close()
    return [dict(row) for row in rows]


def _write_backup_folder(
    zf: zipfile.ZipFile, source: Path, archive_root: str, expected_sha: str,
    display_name: str, reserved: set[str],
) -> None:
    """将已登记的文件夹证据写入备份，并对比登记时的整体摘要。"""
    _add_archive_dir(zf, archive_root, reserved)
    digest_parts: list[str] = []
    for relative, member in _iter_folder_members(source):
        archive_path = f"{archive_root}/{relative}"
        if archive_path in reserved:
            raise ValueError(f"备份附件路径重复：{display_name}")
        reserved.add(archive_path)
        _add_archive_dir(zf, str(PurePosixPath(archive_path).parent), reserved)
        _size, sha = _write_streamed_archive_file(zf, member, archive_path)
        digest_parts.append(f"{relative}\t{sha}")
    actual_sha = hashlib.sha256("\n".join(sorted(digest_parts)).encode("utf-8")).hexdigest()
    if not expected_sha or actual_sha != expected_sha:
        raise ValueError(f"备份时文件夹证据内容已变化：{display_name}")


def _create_backup_locked(proj: AuditProject) -> dict:
    """在项目写锁已持有时生成可自证的一致性完整备份。"""
    import sqlite3

    from limits import MAX_EXTRACT_TOTAL, human_size

    bak_name = f"{_safe(proj.project_name)}_备份_{_now_ts()}.auditbak"
    # 防覆盖：同秒/同名已存在时自动追加序号（审查 F-04 修复）
    bak_path = _unique_path(proj.root.parent, bak_name)
    tmp_bak = bak_path.with_suffix(bak_path.suffix + ".tmp")
    try:
        with tempfile.TemporaryDirectory(prefix="audit_bak_") as td:
            td_path = Path(td)
            # 1) 一致性数据库快照
            snap_path = td_path / "audit.db"
            src = proj._conn
            dst = sqlite3.connect(str(snap_path))
            try:
                src.backup(dst)
            finally:
                dst.close()
            snapshot_files = _snapshot_backup_files(snap_path)
            source_total = snap_path.stat().st_size
            for row in snapshot_files:
                source = proj.attachment_path(row["rel_path"])
                if row.get("mime") == "folder":
                    if not source.is_dir():
                        raise FileNotFoundError(f"备份时文件夹证据已缺失：{row['orig_name']}")
                    source_total += sum(member.stat().st_size for _, member in _iter_folder_members(source))
                else:
                    if not source.is_file():
                        raise FileNotFoundError(f"备份时附件已缺失：{row['orig_name']}")
                    source_total += source.stat().st_size
            if source_total > MAX_EXTRACT_TOTAL:
                raise ValueError(
                    f"项目备份内容约 {human_size(source_total)}，超过可恢复上限 "
                    f"{human_size(MAX_EXTRACT_TOTAL)}；请拆分或清理附件后再备份"
                )
            if shutil.disk_usage(bak_path.parent).free < source_total:
                raise ValueError("备份目标磁盘剩余空间不足")
            # 2) 只写数据库快照中登记的附件，并在流式写入时重新计算摘要。
            with zipfile.ZipFile(tmp_bak, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(snap_path, "audit.db")
                reserved = {"audit.db"}
                for row in snapshot_files:
                    source = proj.attachment_path(row["rel_path"])
                    archive_path = PurePosixPath(str(row["rel_path"]).replace("\\", "/")).as_posix()
                    if archive_path in reserved:
                        raise ValueError(f"备份附件路径重复：{row['orig_name']}")
                    if row.get("mime") == "folder":
                        _write_backup_folder(
                            zf, source, archive_path, str(row.get("sha256") or ""),
                            str(row.get("orig_name") or source.name), reserved,
                        )
                    else:
                        reserved.add(archive_path)
                        _add_archive_dir(zf, str(PurePosixPath(archive_path).parent), reserved)
                        _size, actual_sha = _write_streamed_archive_file(zf, source, archive_path)
                        expected_sha = str(row.get("sha256") or "")
                        if not expected_sha or actual_sha != expected_sha:
                            raise ValueError(f"备份时附件内容已变化：{row['orig_name']}")
            db_size = snap_path.stat().st_size
        os.replace(tmp_bak, bak_path)
    except Exception:
        tmp_bak.unlink(missing_ok=True)
        raise
    return {"filename": bak_path.name, "abs_path": str(bak_path),
            "db_size": db_size}


AUTO_BACKUP_STORE_NAME = "审迹自动备份"


def _copy_object_once(source: Path, objects_dir: Path, sha256: str) -> tuple[bool, int]:
    """将一个证据对象写入内容寻址库；已存在相同摘要时不重复复制。"""
    target = objects_dir / sha256
    if target.is_file():
        return False, target.stat().st_size
    objects_dir.mkdir(parents=True, exist_ok=True)
    temporary = objects_dir / f".{sha256}.{os.getpid()}.tmp"
    try:
        shutil.copyfile(source, temporary)
        if _sha256_of_file(temporary) != sha256:
            raise ValueError(f"附件复制后摘要不一致：{source.name}")
        try:
            os.replace(temporary, target)
            return True, target.stat().st_size
        except FileExistsError:  # 并行任务已写入同一对象，直接复用即可。
            return False, target.stat().st_size
    finally:
        temporary.unlink(missing_ok=True)


def _auto_backup_store(project_uuid: str, target_dir: str | Path) -> Path:
    return Path(target_dir).expanduser().resolve() / AUTO_BACKUP_STORE_NAME / project_uuid


def _iter_folder_members(folder: Path) -> list[tuple[str, Path]]:
    """列出文件夹证据成员；不跟随符号链接，只跳过系统元数据。"""
    members: list[tuple[str, Path]] = []
    for item in folder.rglob("*"):
        if item.is_symlink():
            raise ValueError(f"文件夹证据包含符号链接：{item.name}")
        if not item.is_file():
            continue
        rel = item.relative_to(folder).as_posix()
        if _is_system_metadata_member(rel):
            continue
        members.append((rel, item))
    return sorted(members, key=lambda item: item[0])


def _store_size(path: Path) -> int:
    return sum(item.stat().st_size for item in path.rglob("*") if item.is_file()) if path.exists() else 0


def _manifest_object_hashes(manifest: dict) -> set[str]:
    hashes: set[str] = set()
    for item in manifest.get("attachments", []):
        if item.get("kind") == "folder":
            hashes.update(member.get("sha256", "") for member in item.get("members", []))
        else:
            hashes.add(str(item.get("sha256", "")))
    hashes.discard("")
    return hashes


def _prune_auto_backup_store(store: Path, *, retention_days: int, max_bytes: int) -> dict:
    """保留最近恢复点，随后清理没有被任何恢复点引用的对象。

    调用方必须保证进入本函数时 store 大小不超 max_bytes（新恢复点已在上限
    预检中计入）；本函数只做保留期清理与（防御性的）最旧点裁剪，绝不出现
    "先删除历史恢复点、再报告失败"的破坏路径（S1 修复）。
    """
    points_dir = store / "recovery-points"
    now = datetime.now().timestamp()
    cutoff = now - retention_days * 24 * 60 * 60
    points = sorted((item for item in points_dir.iterdir() if item.is_dir()), key=lambda item: item.name) if points_dir.exists() else []
    removed_points = 0
    for point in points:
        if point.stat().st_mtime < cutoff:
            logger.warning("自动备份裁剪：删除过期恢复点 %s（超过保留期 %d 天）", point.name, retention_days)
            shutil.rmtree(point, ignore_errors=True)
            removed_points += 1
    points = sorted((item for item in points_dir.iterdir() if item.is_dir()), key=lambda item: item.name) if points_dir.exists() else []
    # 至少保留刚生成的最新恢复点；空间不足时删除最早的旧点。
    while len(points) > 1 and _store_size(store) > max_bytes:
        victim = points.pop(0)
        from limits import human_size
        logger.warning("自动备份裁剪：删除最旧恢复点 %s 以满足空间上限 %s", victim.name, human_size(max_bytes))
        shutil.rmtree(victim, ignore_errors=True)
        removed_points += 1

    referenced: set[str] = set()
    all_manifests_valid = True
    for point in points:
        try:
            referenced.update(_manifest_object_hashes(json.loads((point / "manifest.json").read_text(encoding="utf-8"))))
        except (OSError, json.JSONDecodeError):
            # 损坏恢复点不能证明对象仍被引用；保守地保留对象，由后续人工处理。
            all_manifests_valid = False
            break
    objects_dir = store / "objects"
    removed_objects = 0
    if all_manifests_valid and objects_dir.exists():
        for obj in objects_dir.iterdir():
            if obj.is_file() and obj.name not in referenced:
                obj.unlink(missing_ok=True)
                removed_objects += 1
    return {"removed_points": removed_points, "removed_objects": removed_objects, "store_size": _store_size(store)}


def create_incremental_recovery_point(
    proj: AuditProject, *, target_dir: str | Path, max_bytes: int, retention_days: int = 7,
    progress=None, cancelled=None,
) -> dict:
    """创建自动备份恢复点。

    附件使用 SHA-256 内容寻址对象库，已有对象不复制；每次恢复点仅写入一致性
    SQLite 快照和 manifest，因此无变化的 50GB 附件不会在 6 小时后再次复制。
    创建期间持有项目写锁，附件清单、对象摘要和数据库快照属于同一业务时点。
    """
    with proj._lock:
        return _create_incremental_recovery_point_locked(
            proj, target_dir=target_dir, max_bytes=max_bytes, retention_days=retention_days,
            progress=progress, cancelled=cancelled,
        )


def _create_incremental_recovery_point_locked(
    proj: AuditProject, *, target_dir: str | Path, max_bytes: int, retention_days: int = 7,
    progress=None, cancelled=None,
) -> dict:
    """在项目写锁已持有时生成增量恢复点。"""
    import sqlite3

    from limits import human_size

    store = _auto_backup_store(proj.project_uuid, target_dir)
    if store == proj.root.resolve() or store.is_relative_to(proj.root.resolve()):
        raise ValueError("自动备份目录不能位于当前项目内")
    store.mkdir(parents=True, exist_ok=True)
    objects_dir = store / "objects"
    points_dir = store / "recovery-points"
    points_dir.mkdir(parents=True, exist_ok=True)

    rows = [dict(row) for row in proj._conn.execute("SELECT * FROM files ORDER BY id").fetchall()]
    attachment_entries: list[dict] = []
    copy_tasks: list[tuple[Path, str]] = []
    logical_attachment_bytes = 0
    for file_row in rows:
        if cancelled and cancelled():
            raise RuntimeError("自动备份已取消")
        source = proj.attachment_path(file_row["rel_path"])
        if file_row.get("mime") == "folder":
            if not source.is_dir():
                raise ValueError(f"文件夹证据缺失：{file_row['orig_name']}")
            members = []
            for relative, member in _iter_folder_members(source):
                sha = _sha256_of_file(member)
                members.append({"path": relative, "sha256": sha, "size": member.stat().st_size})
                copy_tasks.append((member, sha))
                logical_attachment_bytes += member.stat().st_size
            attachment_entries.append({
                "file_uuid": file_row.get("file_uuid", ""), "rel_path": file_row["rel_path"],
                "kind": "folder", "members": members,
            })
        else:
            if not source.is_file():
                raise ValueError(f"附件缺失：{file_row['orig_name']}")
            sha = _sha256_of_file(source)
            attachment_entries.append({
                "file_uuid": file_row.get("file_uuid", ""), "rel_path": file_row["rel_path"],
                "kind": "file", "sha256": sha, "size": source.stat().st_size,
            })
            copy_tasks.append((source, sha))
            logical_attachment_bytes += source.stat().st_size

    unique_tasks = {sha: source for source, sha in copy_tasks}
    estimated_new = sum(source.stat().st_size for sha, source in unique_tasks.items() if not (objects_dir / sha).is_file())
    estimated_new += proj.db_path.stat().st_size if proj.db_path.exists() else 0
    if shutil.disk_usage(store).free < estimated_new:
        raise ValueError("自动备份目标磁盘剩余空间不足，未创建恢复点")
    # S1 修复：空间上限预检必须在提交新恢复点之前完成——新点加入后 store 即超限时
    # 直接拒绝且不删除任何旧恢复点，杜绝"先删历史恢复点、再报失败"的数据破坏路径。
    estimated_store = _store_size(store) + estimated_new
    if estimated_store > max_bytes:
        raise ValueError(
            f"自动备份将超过最大保留空间 {human_size(max_bytes)}（预计 {human_size(estimated_store)}），"
            "未创建恢复点且未删除任何旧恢复点，请调大自动备份空间上限后重试"
        )

    copied_objects = reused_objects = copied_bytes = 0
    total = len(unique_tasks)
    for index, (sha, source) in enumerate(sorted(unique_tasks.items()), 1):
        if cancelled and cancelled():
            raise RuntimeError("自动备份已取消")
        created, size = _copy_object_once(source, objects_dir, sha)
        if created:
            copied_objects += 1
            copied_bytes += size
        else:
            reused_objects += 1
        if progress:
            progress(index, total, "objects")

    point_name = _now_ts()
    suffix = 1
    while (points_dir / point_name).exists() or (points_dir / f".{point_name}.tmp").exists():
        point_name = f"{_now_ts()}_{suffix}"
        suffix += 1
    stage = points_dir / f".{point_name}.tmp"
    point = points_dir / point_name
    try:
        stage.mkdir(parents=True, exist_ok=False)
        snapshot = stage / "audit.db"
        destination = sqlite3.connect(str(snapshot))
        try:
            proj._conn.backup(destination)
        finally:
            destination.close()
        manifest = {
            "format": 1,
            "project_uuid": proj.project_uuid,
            "project_name": proj.project_name,
            "created_at": _now(),
            "attachments": attachment_entries,
            "logical_attachment_bytes": logical_attachment_bytes,
        }
        (stage / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        os.replace(stage, point)
    except Exception:
        shutil.rmtree(stage, ignore_errors=True)
        raise

    # 空间上限已在上方提交前预检通过（S1 修复），此处裁剪不会再出现
    # "先删旧点、后报失败"；若仍超限（防御：并发写入 store），保留最新点并明确报错。
    prune = _prune_auto_backup_store(store, retention_days=int(retention_days), max_bytes=max_bytes)
    if prune["store_size"] > max_bytes:
        logger.error("自动备份：裁剪后 store 仍超上限 %s（实际 %s），未删除更多旧点，请人工检查",
                     human_size(max_bytes), human_size(prune["store_size"]))
        raise ValueError("自动备份最大保留空间不足且无法通过裁剪满足，请调大上限后重试")
    return {
        "store": str(store), "recovery_point": point.name,
        "copied_objects": copied_objects, "reused_objects": reused_objects,
        "copied_bytes": copied_bytes, **prune,
    }


def list_incremental_recovery_points(project_uuid: str, target_dir: str | Path) -> list[dict]:
    """列出指定项目的可用恢复点，忽略未完整写入的临时目录。"""
    points_dir = _auto_backup_store(project_uuid, target_dir) / "recovery-points"
    if not points_dir.is_dir():
        return []
    result = []
    for point in sorted((item for item in points_dir.iterdir() if item.is_dir() and not item.name.startswith(".")), reverse=True):
        try:
            manifest = json.loads((point / "manifest.json").read_text(encoding="utf-8"))
            if manifest.get("project_uuid") != project_uuid or not (point / "audit.db").is_file():
                continue
            objects_dir = points_dir.parent / "objects"
            missing_objects = 0
            for sha in _manifest_object_hashes(manifest):
                if not (objects_dir / sha).is_file():
                    missing_objects += 1
            result.append({
                "id": point.name, "created_at": manifest.get("created_at", ""),
                "attachments": len(manifest.get("attachments", [])),
                "size": _store_size(point),
                "logical_bytes": int(manifest.get("logical_attachment_bytes") or 0),
                "health": "正常" if not missing_objects else f"缺少 {missing_objects} 个对象",
            })
        except (OSError, json.JSONDecodeError):
            continue
    return result


def _safe_incremental_member_path(folder: Path, raw: str) -> Path:
    relative = PurePosixPath(str(raw or "").replace("\\", "/"))
    if relative.is_absolute() or not relative.parts or any(part in {".", ".."} for part in relative.parts):
        raise ValueError("恢复点包含非法文件夹成员路径")
    candidate = (folder / relative).resolve()
    root = folder.resolve()
    if candidate == root or not candidate.is_relative_to(root):
        raise ValueError("恢复点包含非法文件夹成员路径")
    return candidate


def restore_incremental_recovery_point(
    *, project_uuid: str, backup_target_dir: str | Path, recovery_point_id: str, target_dir: str | Path,
) -> dict:
    """从增量对象库恢复独立项目目录，不覆盖当前或已有项目。"""
    point_id = str(recovery_point_id or "").strip()
    if not point_id or Path(point_id).name != point_id or point_id.startswith("."):
        raise ValueError("恢复点标识非法")
    store = _auto_backup_store(project_uuid, backup_target_dir)
    point = store / "recovery-points" / point_id
    try:
        manifest = json.loads((point / "manifest.json").read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        raise ValueError("恢复点不存在或清单已损坏") from e
    if manifest.get("format") != 1 or manifest.get("project_uuid") != project_uuid:
        raise ValueError("恢复点不属于当前项目")
    snapshot = point / "audit.db"
    if not snapshot.is_file():
        raise ValueError("恢复点缺少数据库快照")

    target = Path(target_dir).expanduser()
    final = target if target.name.endswith(PROJECT_EXT) else target.with_name(target.name + PROJECT_EXT)
    target_existed = target.exists()
    if target_existed and (not target.is_dir() or any(target.iterdir())):
        raise ValueError("恢复目标必须是空文件夹或不存在")
    if final != target and final.exists():
        raise ValueError("恢复目标已存在同名项目目录")

    files_to_restore: list[tuple[Path, str]] = []
    for item in manifest.get("attachments", []):
        rel_path = str(item.get("rel_path") or "")
        if item.get("kind") == "folder":
            for member in item.get("members", []):
                files_to_restore.append((PurePosixPath(rel_path) / PurePosixPath(str(member.get("path") or "")), str(member.get("sha256") or "")))
        else:
            files_to_restore.append((PurePosixPath(rel_path), str(item.get("sha256") or "")))
    if any(not sha for _, sha in files_to_restore):
        raise ValueError("恢复点附件清单缺少摘要")
    required = snapshot.stat().st_size
    objects_dir = store / "objects"
    for _, sha in files_to_restore:
        obj = objects_dir / sha
        if not obj.is_file() or _sha256_of_file(obj) != sha:
            raise ValueError(f"恢复点对象缺失或损坏：{sha[:12]}")
        required += obj.stat().st_size
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        if shutil.disk_usage(target.parent).free < required:
            raise ValueError("恢复目标磁盘剩余空间不足")
        stage = Path(tempfile.mkdtemp(prefix=".audit_auto_restore_", dir=target.parent))
    except OSError as e:
        raise ValueError(f"无法准备恢复目录：{e}") from e
    try:
        shutil.copyfile(snapshot, stage / "audit.db")
        # 先校验数据库，再落附件，防止损坏快照留下部分项目。
        import sqlite3

        connection = sqlite3.connect(str(stage / "audit.db"))
        try:
            if connection.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
                raise ValueError("恢复点数据库完整性校验未通过")
        finally:
            connection.close()
        for rel_path, sha in files_to_restore:
            relative = PurePosixPath(rel_path)
            if relative.is_absolute() or not relative.parts or relative.parts[0] != ATTACH_DIR or any(part in {".", ".."} for part in relative.parts):
                raise ValueError("恢复点包含非法附件路径")
            destination = (stage / relative).resolve()
            attachment_root = (stage / ATTACH_DIR).resolve()
            if not destination.is_relative_to(attachment_root):
                raise ValueError("恢复点包含非法附件路径")
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(objects_dir / sha, destination)
        if target_existed:
            target.rmdir()
        os.replace(stage, final)
    except Exception:
        shutil.rmtree(stage, ignore_errors=True)
        if target_existed and not target.exists():
            target.mkdir(exist_ok=True)
        raise
    harden_project(final)
    return {"path": str(final), "project_name": str(manifest.get("project_name") or final.name)}


def restore_backup(bak_path, target_dir) -> dict:
    """恢复备份：解包 .auditbak 到目标目录（必须为空或不存在）。

    原子性（审查 F-04 修复）：
    - 先解包到临时目录，校验 ZIP 完整、路径安全、audit.db 存在且 integrity_check 通过
    - 全部校验通过后才创建目标目录并迁移内容；失败不留下半成品目标目录
    目录伪装（V3.2）：恢复目标与新建项目一致，自动追加 .auditproj 后缀并隐藏；
    用户选择已有空文件夹时消费该空壳，不留残留。
    """
    import sqlite3

    bak = Path(bak_path)
    target = Path(target_dir).expanduser()
    if not bak.is_file():
        raise ValueError(f"备份文件不存在：{bak}")
    # 目录伪装：恢复目标自动追加 .auditproj 后缀（已带则不重复加），
    # 落位后 harden_project 隐藏，与新建项目行为一致。
    final = target
    if target.name and not target.name.endswith(PROJECT_EXT):
        final = target.with_name(target.name + PROJECT_EXT)
    target_existed = target.exists()
    if target_existed:
        if not target.is_dir():
            raise ValueError(f"恢复目标不是文件夹：{target}")
        if any(target.iterdir()):
            raise ValueError(f"目标目录非空，请选择空文件夹或新目录：{target}")
    if final != target and final.exists():
        raise ValueError(f"恢复目标已存在同名项目目录：{final}。请更换目标目录")
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        # 临时目录放在目标同一文件系统，校验通过后用 os.replace 原子落位。
        stage = Path(tempfile.mkdtemp(prefix=".audit_restore_", dir=target.parent))
    except OSError as e:
        raise ValueError(f"无法准备恢复目录：{e}") from e
    try:
        try:
            with zipfile.ZipFile(bak) as zf:
                bad = zf.testzip()
                if bad:
                    raise ValueError(f"备份文件损坏：{bad}")
                # 防解压炸弹：同时限制成员数与解压总量。
                _validate_archive_size(zf, "备份包")
                # 防目录穿越：校验所有成员名（恶意包可用 ../ 逃逸目标目录）
                target_res = stage.resolve()
                for name in zf.namelist():
                    member = (target_res / name).resolve()
                    if member != target_res and not member.is_relative_to(target_res):
                        raise ValueError(f"备份包包含非法路径：{name}")
                needed = sum(m.file_size for m in zf.infolist())
                _ensure_disk_space(stage, needed)
                zf.extractall(stage)
        except zipfile.BadZipFile:
            raise ValueError("文件不是有效的备份包（.auditbak）")
        if not (stage / "audit.db").exists():
            raise ValueError("备份包缺少 audit.db，恢复失败")
        # 校验数据库完整性、附件相对路径及附件摘要。files.rel_path 来自不可信
        # 备份，必须在移动到目标目录之前拒绝越界或缺失/替换的证据，确保失败
        # 不落半成品项目。
        try:
            conn = sqlite3.connect(str(stage / "audit.db"))
            try:
                status = conn.execute("PRAGMA integrity_check").fetchone()[0]
                if status == "ok":
                    _verify_backup_attachment_snapshot(stage, conn)
            finally:
                conn.close()
        except sqlite3.DatabaseError as e:
            raise ValueError(f"备份数据库无法读取：{e}")
        if status != "ok":
            raise ValueError(f"备份数据库完整性校验未通过（{status}），拒绝恢复")
        # 校验全部通过 → 原子迁移到 final；已有空目录仅在最后一刻移除。
        if target_existed:
            target.rmdir()
        try:
            os.replace(stage, final)
        except OSError:
            if target_existed:
                target.mkdir(exist_ok=True)
            raise
    except ValueError:
        raise
    except OSError as e:
        raise ValueError(f"恢复备份失败：{e}。请检查目标目录权限和剩余空间") from e
    finally:
        if stage.exists():
            shutil.rmtree(stage, ignore_errors=True)
    harden_project(final)  # 隐藏目录（失败不阻断，与新建项目一致）
    return {"path": str(final), "project_name": final.name}


# ═══════════════════════════════════════════════
# 导入问题汇总（Excel）
# ═══════════════════════════════════════════════

IMPORT_HEADERS = ["被审计单位*", "所属版块*", "问题分类", "缺陷定性*", "缺陷描述",
                  "问题金额", "制度依据", "审计建议", "编写人", "审核人"]


def build_import_template(path):
    """生成导入模板 xlsx：Sheet1 表头，Sheet2 填写说明。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(IMPORT_HEADERS)
    fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = [16, 14, 14, 18, 46, 12, 32, 32, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.freeze_panes = "A2"
    # 示例行（灰色，用户可整行删除）
    ws.append(["华电集团XX电厂", "营销管理", "经营管理", "电费回收不及时",
               "示例：某用户电费长期拖欠未回收，金额约120万元。", "120万",
               "《供电营业规则》第XX条", "建议加强催收并建立台账。", "张三", "李四"])
    for cell in ws[2]:
        cell.font = Font(color="808080", italic=True)

    ws2 = wb.create_sheet("填写说明")
    notes = [
        ["问题汇总导入 · 填写说明"],
        [""],
        ["1. 在【导入模板】表逐行填写，每行一条底稿；带 * 的列为必填"],
        ["   - 被审计单位：填写单位名称，不存在的单位会自动创建"],
        ["   - 所属版块、缺陷定性：必填，与程序中版块预设对应"],
        ["2. 可选列：问题分类 / 缺陷描述 / 问题金额 / 制度依据 / 审计建议 / 编写人 / 审核人"],
        ["3. 序号、附件数由程序自动生成，无需填写"],
        ["4. 示例行（灰色斜体）可删除，也可覆盖为自己的数据"],
        ["5. 导入前请勿修改表头文字，删除列会导致导入失败"],
        ["6. 保存为 .xlsx 后，在程序中【导入问题汇总】选择该文件"],
    ]
    for row in notes:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 70
    wb.save(path)


def _import_from_excel_in_place(proj, file_path, operator):
    """解析 xlsx 并导入底稿。返回 {imported, skipped, new_units, errors}。"""
    from limits import MAX_IMPORT_ROWS
    from openpyxl import load_workbook

    # 只读模式按行流式解析，避免 500 MB 工作簿在校验行数前先占满内存。
    wb = load_workbook(file_path, data_only=True, read_only=True)
    rows_to_import = []
    skipped = 0
    errors = []
    try:
        ws = wb["导入模板"] if "导入模板" in wb.sheetnames else wb.active
        required_headers = {"被审计单位", "所属版块", "缺陷定性"}
        header = {}
        header_row = 0
        # 手工模板的表头在第 1 行，程序导出的归档汇总表在标题下方第 2 行；
        # 扫描前 10 行兼容两者，同时避免把项目标题误当成表头。
        for row_number in range(1, 11):
            candidate = {
                str(cell.value).strip().rstrip("*"): cell.column
                for cell in ws[row_number]
                if cell.value
            }
            if required_headers.issubset(candidate):
                header = candidate
                header_row = row_number
                break
        if not header_row:
            raise ValueError("模板缺少必填列：被审计单位、所属版块、缺陷定性（请使用程序提供的模板）")

        def gv(row, name):
            col = header.get(name)
            if col is None:
                return ""
            value = row[col - 1].value
            return "" if value is None else str(value).strip()

        # 先完成资源上限与行级校验，再写项目库，避免第 10001 行才发现超限时
        # 前 10000 行已经部分导入。
        for row in ws.iter_rows(min_row=header_row + 1):
            if row[0].row - header_row > MAX_IMPORT_ROWS:
                raise ValueError(f"导入行数超过上限 {MAX_IMPORT_ROWS} 行，请拆分后再导入")
            unit_name = gv(row, "被审计单位")
            department = gv(row, "所属版块")
            defect_type = gv(row, "缺陷定性")
            if not unit_name and not department and not defect_type:
                continue
            if not unit_name:
                skipped += 1
                errors.append(f"第{row[0].row}行：缺少被审计单位")
                continue
            if not department or not defect_type:
                skipped += 1
                errors.append(f"第{row[0].row}行：所属版块/缺陷定性为必填")
                continue
            rows_to_import.append({
                "unit_name": unit_name,
                "department": department,
                "category": gv(row, "问题分类"),
                "defect_type": defect_type,
                "defect_desc": gv(row, "缺陷描述"),
                "amount": gv(row, "问题金额"),
                "regulation_basis": gv(row, "制度依据"),
                "suggestion": gv(row, "审计建议"),
                "author": gv(row, "编写人") or operator,
                "reviewer": gv(row, "审核人"),
            })
    finally:
        wb.close()

    unit_cache = {u["name"]: u["id"] for u in proj.list_units()}
    imported = new_units = 0
    for item in rows_to_import:
        unit_name = item.pop("unit_name")
        if unit_name not in unit_cache:
            uid = proj.add_unit(unit_name, operator)
            unit_cache[unit_name] = uid
            new_units += 1
        proj.add_issue(unit_cache[unit_name], operator=operator, **item)
        imported += 1
    proj.log(operator, "导入问题汇总",
             f"成功 {imported} 条，跳过 {skipped} 条，新建单位 {new_units} 个")
    # errors 完整返回（T7：前端可导出完整导入报告）；超大导入最多也就上限行数条，可控
    return {"imported": imported, "skipped": skipped,
            "new_units": new_units, "errors": errors}


def merge_import(proj, zip_paths, operator):
    """合并导入：审计经理汇总多个归档 ZIP 到当前项目。

    每个 ZIP：审计问题汇总.xlsx（建单位/底稿）+ 附件-单位/... 目录（还原附件）。
    附件匹配：单位名 + 缺陷定性（同定性多条按单位内创建顺序依次分配）。
    返回 {imported, skipped, files, new_units, errors}。
    """
    import tempfile
    import zipfile as _zipfile

    unit_cache = {u["name"]: u["id"] for u in proj.list_units()}
    imported = skipped = files = new_units = 0
    errors = []

    for zp in zip_paths:
        tmp = Path(tempfile.mkdtemp(prefix="audit_merge_"))
        try:
            with _zipfile.ZipFile(zp) as zf:
                # 防解压炸弹：同时限制成员数与解压总量。
                _validate_archive_size(zf, "归档包")
                # 找汇总表
                xlsx_name = next((n for n in zf.namelist()
                                  if n.endswith("审计问题汇总.xlsx")), None)
                if not xlsx_name:
                    errors.append(f"{Path(zp).name}：未找到 审计问题汇总.xlsx，跳过")
                    skipped += 1
                    shutil.rmtree(tmp, ignore_errors=True)
                    continue
                # 防目录穿越
                target_res = tmp.resolve()
                for name in zf.namelist():
                    member = (target_res / name).resolve()
                    if member != target_res and not member.is_relative_to(target_res):
                        raise ValueError(f"导入包包含非法路径：{name}")
                needed = sum(m.file_size for m in zf.infolist())
                _ensure_disk_space(tmp, needed)
                zf.extractall(tmp)
        except Exception as e:
            errors.append(f"{Path(zp).name}：{e}")
            skipped += 1
            shutil.rmtree(tmp, ignore_errors=True)
            continue

        try:
            # 1) 建单位/底稿（复用 Excel 导入）
            existing_issue_ids = {
                issue["id"]
                for unit in proj.list_units()
                for issue in proj.list_issues(unit["id"])
            }
            r = import_from_excel(proj, tmp / xlsx_name, operator)
            imported += r["imported"]
            skipped += r["skipped"]
            new_units += r["new_units"]
            errors.extend(r["errors"])
            unit_cache = {u["name"]: u["id"] for u in proj.list_units()}

            # 2) 附件还原：附件-单位/.../序号.定性/文件
            # 归档 ZIP 以“项目名称”为根目录，附件目录与汇总表位于同一级。
            root = (tmp / xlsx_name).parent
            for udir in root.glob("附件-*"):
                unit_name = udir.name[len("附件-"):]
                uid = unit_cache.get(unit_name)
                if uid is None:
                    continue
                # 单位内底稿按创建顺序（= 汇总表行序 = 原序号序）
                # 只匹配本次导入生成的底稿，绝不能把归档证据误挂到目标项目原有的
                # 同名单位/同定性底稿。
                issues = [issue for issue in proj.list_issues(uid)
                          if issue["id"] not in existing_issue_ids]
                used = set()
                for idir in sorted(udir.rglob("*"), key=lambda p: str(p)):
                    if not idir.is_dir():
                        continue
                    m = re.match(r"(\d+)\.(.+)", idir.name)
                    if not m or m.group(2) in ("审计问题汇总",):
                        continue
                    defect = m.group(2)
                    # 匹配：同单位同定性、尚未分配附件的底稿
                    cand = [i for i in issues
                            if i["defect_type"] == defect and i["id"] not in used]
                    if not cand:
                        continue
                    iss = cand[0]
                    used.add(iss["id"])
                    for fp in sorted(idir.iterdir()):
                        if fp.name in SYSTEM_METADATA_NAMES:
                            continue
                        try:
                            if fp.is_dir():
                                members = [
                                    (member.relative_to(fp).as_posix(), str(member))
                                    for member in sorted(fp.rglob("*"))
                                    if member.is_file() and not _is_system_metadata_member(member.relative_to(fp))
                                ]
                                rec = proj.add_folder(uid, members, fp.name, operator)
                            elif fp.is_file():
                                rec = proj.add_file(uid, str(fp), operator, orig_name=fp.name)
                            else:
                                continue
                            proj.link_file(iss["id"], rec["id"], operator)
                            files += 1
                        except Exception as e:
                            errors.append(f"{unit_name}/{idir.name}/{fp.name}：{e}")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    proj.log(operator, "合并导入",
             f"成功 {imported} 条底稿、附件 {files} 个，跳过 {skipped}，新建单位 {new_units} 个")
    return {"imported": imported, "skipped": skipped, "files": files,
            "new_units": new_units, "errors": errors[:30]}


def _merge_backups_in_place(proj, bak_paths, operator):
    """在隔离项目内执行备份合并。

    这是 ``merge_backups`` 的内部实现。正式入口总是在临时项目中调用本函数，
    只有全部来源成功后才替换当前项目，禁止半批次结果直接进入正式底稿。
    """
    import sqlite3
    import tempfile
    import zipfile as _zipfile

    unit_cache = {u["name"]: u["id"] for u in proj.list_units()}
    # T9：合并前的初始单位集合（检测"同名已存在"必须用它，避免把刚创建的单位误判）
    initial_unit_names = set(unit_cache)
    units = issues = files = folders = depts = versions = requests = source_logs = 0
    errors = []
    # T9 冲突清单：合并行为与 v1.1 一致（复用/重排/都保留/去重合并），
    # 但把每一处"冲突被正确处理"的情况报告出来，供审计经理核对。
    conflicts = []

    for zp in bak_paths:
        tmp = Path(tempfile.mkdtemp(prefix="audit_merge_"))
        try:
            with _zipfile.ZipFile(zp) as zf:
                bad = zf.testzip()
                if bad:
                    errors.append(f"{Path(zp).name}：备份损坏（{bad}）")
                    continue
                # 防解压炸弹：同时限制成员数与解压总量。
                _validate_archive_size(zf, "备份包")
                # 防目录穿越
                target_res = tmp.resolve()
                for name in zf.namelist():
                    member = (target_res / name).resolve()
                    if member != target_res and not member.is_relative_to(target_res):
                        raise ValueError(f"备份包包含非法路径：{name}")
                needed = sum(m.file_size for m in zf.infolist())
                _ensure_disk_space(tmp, needed)
                zf.extractall(tmp)
            db = tmp / "audit.db"
            if not db.exists():
                errors.append(f"{Path(zp).name}：缺少 audit.db，跳过")
                continue

            conn = sqlite3.connect(str(db))
            conn.row_factory = sqlite3.Row
            try:
                # 单位
                b_units = conn.execute("SELECT id, name FROM units ORDER BY sort_order, id").fetchall()
                unit_map = {}
                for bu in b_units:
                    name = bu["name"]
                    if name not in unit_cache:
                        uid = proj.add_unit(name, operator)
                        unit_cache[name] = uid
                        units += 1
                    unit_map[bu["id"]] = unit_cache[name]

                # 底稿
                b_issues = conn.execute(
                    "SELECT * FROM issues ORDER BY unit_id, seq, id").fetchall()
                try:
                    source_versions = conn.execute(
                        "SELECT issue_id, version_no, snapshot, saved_by, created_at "
                        "FROM issue_versions ORDER BY issue_id, version_no"
                    ).fetchall()
                except sqlite3.OperationalError:
                    source_versions = []
                versions_by_issue: dict[int, list[sqlite3.Row]] = {}
                for version in source_versions:
                    versions_by_issue.setdefault(version["issue_id"], []).append(version)
                issue_map = {}
                # T9 冲突检测 1/2：同名单位不同底稿（保留两边）+ 同 seq 底稿（自动重排）
                b_by_unit: dict[int, list] = {}
                for bi in b_issues:
                    b_by_unit.setdefault(bi["unit_id"], []).append(bi)
                for bu in b_units:
                    if bu["name"] not in initial_unit_names:
                        continue  # 全新单位无冲突
                    target_uid = unit_cache[bu["name"]]
                    b_list = b_by_unit.get(bu["id"], [])
                    if b_list:
                        conflicts.append({
                            "type": "unit_exists",
                            "message": f"单位「{bu['name']}」已存在，其 {len(b_list)} 条底稿将并入（保留两边）",
                        })
                    # 同 seq 冲突：备份底稿的 seq 与目标单位已有 seq 重叠 → 自动重排
                    existing_seqs = {iss["seq"] for iss in proj.list_issues(target_uid)}
                    overlap = [bi["seq"] for bi in b_list if bi["seq"] in existing_seqs]
                    if overlap:
                        conflicts.append({
                            "type": "seq_reshuffle",
                            "message": f"单位「{bu['name']}」底稿序号与现有重叠（{sorted(overlap)}），将自动重排",
                        })
                for bi in b_issues:
                    bi = dict(bi)  # 旧备份可没有 v3 的 category 列，统一用 dict.get 兼容
                    if bi["unit_id"] not in unit_map:
                        continue
                    nid = proj.add_issue(
                        unit_map[bi["unit_id"]], operator=operator,
                        department=bi["department"] or "",
                        category=bi.get("category", "") or "",
                        defect_type=bi["defect_type"] or "",
                        defect_desc=bi["defect_desc"] or "", amount=bi["amount"] or "",
                        regulation_basis=bi["regulation_basis"] or "",
                        suggestion=bi["suggestion"] or "",
                        author=bi["author"] or operator, reviewer=bi["reviewer"] or "",
                    )
                    issue_map[bi["id"]] = nid
                    issues += 1
                    source_status = bi.get("status") or AuditProject.STATUS_DRAFT
                    if source_status not in AuditProject.STATUSES:
                        errors.append(
                            f"{Path(zp).name}：底稿 {bi['id']} 的状态“{source_status}”无效，已按草稿导入"
                        )
                        source_status = AuditProject.STATUS_DRAFT
                    imported_versions = versions_by_issue.get(bi["id"], [])
                    current = proj.get_issue(nid)
                    with proj._lock, proj._conn:
                        proj._conn.execute(
                            "UPDATE issues SET status=?, created_at=?, updated_at=? WHERE id=?",
                            (
                                source_status,
                                bi.get("created_at") or current["created_at"],
                                bi.get("updated_at") or current["updated_at"],
                                nid,
                            ),
                        )
                        if imported_versions:
                            # add_issue 自动生成的 v1 仅是导入占位；来源版本存在时用完整
                            # 来源版本链替换，保留保存人、时间和历史快照。
                            proj._conn.execute("DELETE FROM issue_versions WHERE issue_id=?", (nid,))
                            proj._conn.executemany(
                                "INSERT INTO issue_versions(issue_id, version_no, snapshot, saved_by, created_at) "
                                "VALUES(?,?,?,?,?)",
                                [
                                    (nid, version["version_no"], version["snapshot"],
                                     version["saved_by"], version["created_at"])
                                    for version in imported_versions
                                ],
                            )
                            versions += len(imported_versions)

                # 文件（含文件夹实体；关联与未关联证据均需导入）
                b_files = conn.execute("SELECT * FROM files ORDER BY id").fetchall()
                file_map = {}
                # T9 冲突检测 3：同名附件不同内容（都保留）
                # 目标单位已有附件（orig_name → sha256），用于识别同名不同内容
                target_names: dict[tuple[int, str], str] = {}
                for fu in proj.list_units():
                    for f in proj.list_files(fu["id"]):
                        if f.get("mime") != "folder":
                            target_names[(fu["id"], f.get("orig_name", ""))] = f.get("sha256", "")
                for bf in b_files:
                    if bf["unit_id"] not in unit_map:
                        continue
                    bf = dict(bf)  # sqlite3.Row 无 .get()，转 dict（T9 冲突检测需要）
                    src = _extracted_attachment_path(tmp, bf["rel_path"])
                    if bf["mime"] == "folder" and src.is_dir():
                        rec = _import_folder_dir(proj, unit_map[bf["unit_id"]],
                                                 src, bf["orig_name"], operator)
                        if rec:
                            file_map[bf["id"]] = rec["id"]
                            folders += 1
                        else:
                            errors.append(f"{Path(zp).name}：文件夹附件 {bf['orig_name']} 导入失败")
                    elif src.is_file():
                        # 同名但内容（sha256）不同 → 冲突：两个都保留（stored_name 用 uuid 不冲突）
                        key = (unit_map[bf["unit_id"]], bf.get("orig_name", ""))
                        if key in target_names and target_names[key] != bf.get("sha256", ""):
                            conflicts.append({
                                "type": "file_same_name",
                                "message": f"附件「{bf['orig_name']}」与现有附件同名但内容不同，将同时保留",
                            })
                        rec = proj.add_file(unit_map[bf["unit_id"]], str(src), operator,
                                            orig_name=bf["orig_name"])
                        file_map[bf["id"]] = rec["id"]
                        files += 1
                    else:
                        errors.append(
                            f"{Path(zp).name}：附件实体缺失或类型不匹配：{bf.get('orig_name', bf['id'])}"
                        )

                # 关联
                for link in conn.execute("SELECT issue_id, file_id FROM issue_files").fetchall():
                    if link["issue_id"] in issue_map and link["file_id"] in file_map:
                        try:
                            proj.link_file(issue_map[link["issue_id"]],
                                           file_map[link["file_id"]], operator)
                        except (KeyError, ValueError) as exc:
                            errors.append(
                                f"{Path(zp).name}：附件关联 {link['issue_id']}/{link['file_id']} 导入失败：{exc}"
                            )
                # 独占证据语义随备份迁移：目标附件只能保留来源指定底稿这一条关联。
                for bf in b_files:
                    bf = dict(bf)
                    exclusive_to = bf.get("exclusive_to")
                    if not exclusive_to:
                        continue
                    if bf["id"] not in file_map or exclusive_to not in issue_map:
                        errors.append(
                            f"{Path(zp).name}：附件 {bf['id']} 的独占底稿不存在，已按普通附件导入"
                        )
                        continue
                    try:
                        proj.link_file_exclusive(
                            issue_map[exclusive_to], file_map[bf["id"]], operator
                        )
                    except (KeyError, ValueError) as exc:
                        errors.append(f"{Path(zp).name}：附件 {bf['id']} 独占关系导入失败：{exc}")

                # 项目级资料请求与底稿、单位、附件一同迁移。来源为旧版本时表不存在，
                # 跳过即可；来源 UUID 与当前项目冲突则生成新的本地 UUID，避免覆盖历史。
                try:
                    source_requests = conn.execute("SELECT * FROM project_requests ORDER BY created_at, request_uuid").fetchall()
                except sqlite3.OperationalError:
                    source_requests = []
                for source_request in source_requests:
                    request = dict(source_request)
                    request_uuid = str(request.get("request_uuid") or uuid.uuid4())
                    if proj._conn.execute(
                        "SELECT 1 FROM project_requests WHERE request_uuid=?", (request_uuid,)
                    ).fetchone():
                        request_uuid = str(uuid.uuid4())
                    status = str(request.get("status") or "open")
                    if status not in {"open", "provided", "verified", "withdrawn"}:
                        errors.append(f"{Path(zp).name}：资料请求 {request.get('title', '')} 状态无效，已按待提供导入")
                        status = "open"
                    with proj._lock, proj._conn:
                        proj._conn.execute(
                            "INSERT INTO project_requests(request_uuid,unit_id,issue_id,title,detail,responsible,due_date,status,provided_file_id,note,created_by,created_at,updated_by,updated_at) "
                            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (
                                request_uuid,
                                unit_map.get(request.get("unit_id")) if request.get("unit_id") is not None else None,
                                issue_map.get(request.get("issue_id")) if request.get("issue_id") is not None else None,
                                str(request.get("title") or "未命名资料请求"), str(request.get("detail") or ""),
                                str(request.get("responsible") or ""), str(request.get("due_date") or ""), status,
                                file_map.get(request.get("provided_file_id")) if request.get("provided_file_id") is not None else None,
                                str(request.get("note") or ""), str(request.get("created_by") or operator),
                                str(request.get("created_at") or _now()), str(request.get("updated_by") or operator),
                                str(request.get("updated_at") or request.get("created_at") or _now()),
                            ),
                        )
                    requests += 1

                # 版块预设合并
                try:
                    meta = conn.execute("SELECT value FROM meta WHERE key='departments'").fetchone()
                    if meta and meta["value"]:
                        cur_depts = json.loads(proj.get_meta("departments", "[]"))
                        new_depts = []
                        for d in json.loads(meta["value"]):
                            if d and d not in cur_depts:
                                cur_depts.append(d)
                                new_depts.append(d)
                                depts += 1
                        if new_depts:
                            # T9 冲突检测 4：版块预设差异（去重合并）
                            conflicts.append({
                                "type": "dept_merge",
                                "message": f"新增版块预设 {len(new_depts)} 个：{'、'.join(new_depts)}",
                            })
                        proj.set_meta("departments", json.dumps(cur_depts, ensure_ascii=False))
                except (ValueError, TypeError, json.JSONDecodeError) as exc:
                    errors.append(f"{Path(zp).name}：版块预设导入失败：{exc}")
                # 问题分类预设合并（v3 新增；旧备份无该键时自然跳过）。
                try:
                    meta = conn.execute("SELECT value FROM meta WHERE key='categories'").fetchone()
                    if meta and meta["value"]:
                        current_categories = json.loads(proj.get_meta("categories", "[]"))
                        new_categories = []
                        for category in json.loads(meta["value"]):
                            if category and category not in current_categories:
                                current_categories.append(category)
                                new_categories.append(category)
                        if new_categories:
                            conflicts.append({
                                "type": "category_merge",
                                "message": f"新增问题分类预设 {len(new_categories)} 个：{'、'.join(new_categories)}",
                            })
                        proj.set_meta("categories", json.dumps(current_categories, ensure_ascii=False))
                except (ValueError, TypeError, json.JSONDecodeError) as exc:
                    errors.append(f"{Path(zp).name}：问题分类预设导入失败：{exc}")

                # 合并不能只保留最终数据：来源项目的操作日志也要以追加记录方式
                # 留在汇总项目中。目标项目会重建自己的哈希链，原日志的事件 UUID、
                # 原操作者、时间和链哈希则作为不可变来源凭据一并保留。
                try:
                    rows = conn.execute(
                        "SELECT event_uuid, operator, action, target, detail, created_at, "
                        "actor_account, actor_uid, device_id, prev_hash, event_hash "
                        "FROM audit_log ORDER BY id"
                    ).fetchall()
                    source_uuid_row = conn.execute(
                        "SELECT value FROM meta WHERE key='project_uuid'"
                    ).fetchone()
                    source_uuid = str(source_uuid_row["value"] if source_uuid_row else "")
                    for row in rows:
                        source = dict(row)
                        detail = json.dumps({
                            "source_project_uuid": source_uuid,
                            "source_event_uuid": source.get("event_uuid", ""),
                            "source_operator": source.get("operator", ""),
                            "source_account": source.get("actor_account", ""),
                            "source_uid": source.get("actor_uid", ""),
                            "source_device": source.get("device_id", ""),
                            "source_created_at": source.get("created_at", ""),
                            "source_detail": source.get("detail", ""),
                            "source_prev_hash": source.get("prev_hash", ""),
                            "source_event_hash": source.get("event_hash", ""),
                        }, ensure_ascii=False, separators=(",", ":"))
                        proj.log(
                            operator,
                            "保留来源操作日志",
                            f"{Path(zp).name} · {source.get('action', '')} · {source.get('target', '')}",
                            detail,
                        )
                        source_logs += 1
                except sqlite3.OperationalError:
                    # v1.1 之前的备份可能还没有 audit_log；可兼容导入，但明确留痕。
                    proj.log(operator, "来源日志缺失", Path(zp).name, "来源备份不含操作日志表")
            finally:
                conn.close()
        except Exception as e:
            errors.append(f"{Path(zp).name}：{e}")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    batch_uuid = proj.record_merge_batch(operator, [str(Path(path).resolve()) for path in bak_paths], conflicts)
    proj.log(operator, "合并导入备份",
             f"单位 {units} 个、底稿 {issues} 条、版本 {versions} 个、附件 {files} 个、文件夹 {folders} 个、资料请求 {requests} 条、版块预设 {depts} 个"
             + (f"、来源日志 {source_logs} 条" if source_logs else "")
             + (f"、冲突 {len(conflicts)} 处" if conflicts else "")
             + f"；批次 {batch_uuid}")
    return {"units": units, "issues": issues, "files": files, "requests": requests,
            "folders": folders, "depts": depts, "versions": versions, "errors": errors[:30],
            "conflicts": conflicts, "source_logs": source_logs, "batch_uuid": batch_uuid}


def _link_or_copy(source: str, target: str) -> str:
    """克隆暂存附件时优先硬链接，避免为 50GB 项目复制完整附件库。"""
    try:
        os.link(source, target)
    except OSError:
        shutil.copy2(source, target)
    return target


def _clone_project_for_merge(proj: AuditProject, stage_root: Path) -> AuditProject:
    """创建当前项目的隔离合并副本；数据库一致性快照 + 附件硬链接克隆。"""
    import sqlite3

    stage_root.mkdir(parents=True, exist_ok=False)
    stage_db = stage_root / "audit.db"
    with proj._lock:
        backup_conn = sqlite3.connect(stage_db)
        try:
            proj._conn.backup(backup_conn)
        finally:
            backup_conn.close()
    source_attachments = proj.root / ATTACH_DIR
    if source_attachments.exists():
        shutil.copytree(source_attachments, stage_root / ATTACH_DIR, copy_function=_link_or_copy)
    stage = AuditProject(stage_root)
    stage.set_audit_identity(proj._actor_uid, proj._device_id)
    return stage


def _replace_project_from_merge_stage(proj: AuditProject, stage: AuditProject) -> None:
    """用已验证的暂存项目替换正式数据库和附件库，并在失败时恢复原项目。"""
    import sqlite3

    stage_root = stage.root
    stage.close()
    target_db = proj.db_path
    target_attachments = proj.root / ATTACH_DIR
    stage_db = stage_root / "audit.db"
    stage_attachments = stage_root / ATTACH_DIR
    rollback_db = stage_root / "rollback_audit.db"
    rollback_attachments = stage_root / "rollback_attachments"
    moved_old_db = installed_stage_db = moved_old_attachments = installed_stage_attachments = False

    with proj._lock:
        proj._swapping = True
        proj._conn.close()
        try:
            os.replace(target_db, rollback_db)
            moved_old_db = True
            os.replace(stage_db, target_db)
            installed_stage_db = True
            os.replace(target_attachments, rollback_attachments)
            moved_old_attachments = True
            os.replace(stage_attachments, target_attachments)
            installed_stage_attachments = True
        except Exception as e:
            # 文件替换由同一文件系统内的 rename 完成；任一步失败都按相反顺序恢复。
            if installed_stage_attachments and target_attachments.exists():
                os.replace(target_attachments, stage_attachments)
            if moved_old_attachments and rollback_attachments.exists():
                os.replace(rollback_attachments, target_attachments)
            if installed_stage_db and target_db.exists():
                os.replace(target_db, stage_db)
            if moved_old_db and rollback_db.exists():
                os.replace(rollback_db, target_db)
            # I2 修复：Windows 下附件目录被其他程序占用（如预览窗口）时给出可执行提示
            if isinstance(e, PermissionError):
                raise PermissionError(
                    "合并/导入提交失败：附件文件正被其他程序占用（例如 PDF 预览窗口）。"
                    "请关闭项目中已打开的文件后重试"
                ) from e
            raise
        finally:
            proj._swapping = False
            proj._conn = sqlite3.connect(proj.db_path, check_same_thread=False)
            proj._conn.row_factory = sqlite3.Row
            proj._conn.execute("PRAGMA foreign_keys = ON")


def merge_backups(proj: AuditProject, bak_paths, operator: str) -> dict:
    """以暂存项目执行整批合并，任一来源失败时当前项目保持完全不变。"""
    stage_root = Path(tempfile.mkdtemp(prefix=".audit_merge_stage_", dir=proj.root.parent))
    shutil.rmtree(stage_root)
    stage: AuditProject | None = None
    try:
        stage = _clone_project_for_merge(proj, stage_root)
        result = _merge_backups_in_place(stage, bak_paths, operator)
        if result["errors"]:
            raise ValueError("合并未提交：" + "；".join(result["errors"][:3]))
        committed_stage = stage
        stage = None  # _replace 会关闭它；失败时也不能在 finally 中重复关闭。
        _replace_project_from_merge_stage(proj, committed_stage)
        return result
    finally:
        if stage is not None:
            stage.close()
        shutil.rmtree(stage_root, ignore_errors=True)


def import_from_excel(proj: AuditProject, file_path, operator: str) -> dict:
    """在隔离副本导入 Excel，全部成功后才替换正式项目。

    行级校验虽然会在写入前完成，但磁盘满、SQLite 错误或未知异常仍可能发生在
    写入中。复用批次合并的暂存替换机制，保证用户不会看到“前半份已导入”。
    """
    stage_root = Path(tempfile.mkdtemp(prefix=".audit_excel_stage_", dir=proj.root.parent))
    shutil.rmtree(stage_root)
    stage: AuditProject | None = None
    try:
        stage = _clone_project_for_merge(proj, stage_root)
        result = _import_from_excel_in_place(stage, file_path, operator)
        committed_stage = stage
        stage = None
        _replace_project_from_merge_stage(proj, committed_stage)
        return result
    finally:
        if stage is not None:
            stage.close()
        shutil.rmtree(stage_root, ignore_errors=True)


def preflight_excel_import(proj: AuditProject, file_path) -> dict:
    """在临时项目副本内执行导入校验，保证预检不会写当前项目。"""
    stage_root = Path(tempfile.mkdtemp(prefix=".audit_excel_preflight_", dir=proj.root.parent))
    shutil.rmtree(stage_root)
    stage: AuditProject | None = None
    try:
        stage = _clone_project_for_merge(proj, stage_root)
        result = _import_from_excel_in_place(stage, file_path, "预检")
        return {
            "imported": result["imported"], "skipped": result["skipped"],
            "new_units": result["new_units"], "errors": result["errors"],
        }
    finally:
        if stage is not None:
            stage.close()
        shutil.rmtree(stage_root, ignore_errors=True)


def merge_preflight(proj: AuditProject, bak_paths: list[str | Path]) -> dict:
    """只读预检备份来源，识别不可执行项与需要负责人确认的并存冲突。

    为适配 50GB 附件，预检仅从 ZIP 流式取出 ``audit.db``，不解压附件库、也不
    向当前项目写数据。实际合并仍会重新校验来源，预检结果不能被当作执行凭据。
    """
    import sqlite3

    target_units = {unit["name"]: unit["id"] for unit in proj.list_units()}
    target_state = {
        "units": [[unit["id"], unit["name"]] for unit in proj.list_units()],
        "issues": [list(row) for row in proj._conn.execute(
            "SELECT issue_uuid, unit_id, seq, updated_at FROM issues WHERE deleted_at IS NULL ORDER BY id"
        ).fetchall()],
        "files": [list(row) for row in proj._conn.execute(
            "SELECT file_uuid, unit_id, sha256, size FROM files ORDER BY id"
        ).fetchall()],
        "last_log": list(proj._conn.execute(
            "SELECT id, event_hash FROM audit_log ORDER BY id DESC LIMIT 1"
        ).fetchone() or (0, "")),
    }
    target_fingerprint = hashlib.sha256(json.dumps(
        target_state, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")).hexdigest()
    target_sequences = {
        unit_id: {issue["seq"] for issue in proj.list_issues(unit_id)}
        for unit_id in target_units.values()
    }
    target_file_hashes: dict[tuple[str, str], set[str]] = {}
    for unit in proj.list_units():
        for file in proj.list_files(unit["id"]):
            if file.get("mime") != "folder":
                target_file_hashes.setdefault((unit["name"], file.get("orig_name", "")), set()).add(
                    str(file.get("sha256") or "")
                )
    try:
        target_departments = set(json.loads(proj.get_meta("departments", "[]")))
    except (TypeError, json.JSONDecodeError):
        target_departments = set()
    try:
        target_categories = set(json.loads(proj.get_meta("categories", "[]")))
    except (TypeError, json.JSONDecodeError):
        target_categories = set()
    blockers: list[dict] = []
    conflicts: list[dict] = []
    sources: list[dict] = []
    seen_project_uuids: set[str] = set()
    fingerprints: list[dict] = []

    for raw in bak_paths:
        path = Path(raw).expanduser()
        label = path.name
        if path.suffix.lower() != ".auditbak" or not path.is_file():
            blockers.append({"source": label, "code": "invalid_path", "message": "备份文件不存在或不是 .auditbak"})
            continue
        stat = path.stat()
        try:
            with tempfile.TemporaryDirectory(prefix="audit_merge_preview_") as td:
                extract_root = Path(td).resolve()
                snapshot = extract_root / "audit.db"
                with zipfile.ZipFile(path) as zf:
                    infos = _validate_archive_size(zf, "备份包")
                    names = [info.filename for info in infos]
                    db_name = next((name for name in names if name == "audit.db"), None)
                    if not db_name:
                        raise ValueError("缺少 audit.db")
                    for name in names:
                        member = (extract_root / name).resolve()
                        if member != extract_root and not member.is_relative_to(extract_root):
                            raise ValueError(f"包含非法路径：{name}")
                    with zf.open(db_name) as source, open(snapshot, "wb") as target:
                        shutil.copyfileobj(source, target, length=1 << 20)
                connection = sqlite3.connect(f"file:{snapshot}?mode=ro", uri=True)
                connection.row_factory = sqlite3.Row
                try:
                    integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
                    if integrity != "ok":
                        raise ValueError("audit.db 完整性校验失败")
                    tables = {row[0] for row in connection.execute(
                        "SELECT name FROM sqlite_master WHERE type='table'"
                    ).fetchall()}
                    if not {"units", "issues", "files"}.issubset(tables):
                        raise ValueError("audit.db 缺少单位、底稿或附件表")
                    try:
                        project_uuid_row = connection.execute(
                            "SELECT value FROM meta WHERE key='project_uuid'"
                        ).fetchone()
                        source_uuid = str(project_uuid_row[0] if project_uuid_row else "")
                    except sqlite3.OperationalError:
                        source_uuid = ""
                    units = connection.execute("SELECT id, name FROM units ORDER BY sort_order, id").fetchall()
                    issues = connection.execute("SELECT unit_id, seq FROM issues ORDER BY unit_id, seq, id").fetchall()
                    source_files = connection.execute(
                        "SELECT unit_id, orig_name, sha256, mime FROM files ORDER BY id"
                    ).fetchall()
                    attachment_count = len(source_files)
                    try:
                        row = connection.execute("SELECT value FROM meta WHERE key='departments'").fetchone()
                        source_departments = set(json.loads(row["value"] if row else "[]"))
                    except (sqlite3.OperationalError, TypeError, json.JSONDecodeError):
                        source_departments = set()
                    try:
                        row = connection.execute("SELECT value FROM meta WHERE key='categories'").fetchone()
                        source_categories = set(json.loads(row["value"] if row else "[]"))
                    except (sqlite3.OperationalError, TypeError, json.JSONDecodeError):
                        source_categories = set()
                finally:
                    connection.close()
        except (OSError, ValueError, zipfile.BadZipFile, sqlite3.Error) as error:
            blockers.append({"source": label, "code": "invalid_backup", "message": str(error)})
            continue

        if source_uuid and source_uuid == proj.project_uuid:
            blockers.append({"source": label, "code": "same_project", "message": "来源备份与当前项目相同，不能合并自身"})
        if source_uuid and source_uuid in seen_project_uuids:
            blockers.append({"source": label, "code": "duplicate_source", "message": "同一来源项目已在本批次中选择，不能重复合并"})
        if source_uuid:
            seen_project_uuids.add(source_uuid)

        units_by_id = {row["id"]: row["name"] for row in units}
        for unit_id, name in units_by_id.items():
            if name not in target_units:
                continue
            source_issue_count = sum(1 for issue in issues if issue["unit_id"] == unit_id)
            if source_issue_count:
                conflicts.append({
                    "source": label, "type": "unit_exists", "resolution": "保留双方",
                    "message": f"单位「{name}」已存在，将并存导入其 {source_issue_count} 条底稿",
                })
            overlaps = sorted({
                issue["seq"] for issue in issues
                if issue["unit_id"] == unit_id and issue["seq"] in target_sequences[target_units[name]]
            })
            if overlaps:
                conflicts.append({
                    "source": label, "type": "seq_reassign", "resolution": "保留双方并重新分配来源编号",
                    "message": f"单位「{name}」有 {len(overlaps)} 个编号冲突（{overlaps[:10]}），来源底稿将保留并重新编号",
                })
        for file in source_files:
            unit_name = units_by_id.get(file["unit_id"])
            if not unit_name or file["mime"] == "folder":
                continue
            existing_hashes = target_file_hashes.get((unit_name, file["orig_name"]), set())
            source_hash = str(file["sha256"] or "")
            if existing_hashes and source_hash not in existing_hashes:
                conflicts.append({
                    "source": label, "type": "file_same_name", "resolution": "保留双方",
                    "message": f"单位「{unit_name}」附件「{file['orig_name']}」同名但内容不同，将同时保留",
                })
        new_departments = sorted(str(item) for item in source_departments - target_departments if str(item).strip())
        if new_departments:
            conflicts.append({
                "source": label, "type": "dept_merge", "resolution": "合并预设，不覆盖现有项目",
                "message": f"将新增 {len(new_departments)} 个版块预设：{'、'.join(new_departments[:10])}",
            })
        new_categories = sorted(str(item) for item in source_categories - target_categories if str(item).strip())
        if new_categories:
            conflicts.append({
                "source": label, "type": "category_merge", "resolution": "合并预设，不覆盖现有项目",
                "message": f"将新增 {len(new_categories)} 个问题分类预设：{'、'.join(new_categories[:10])}",
            })
        source = {
            "name": label, "project_uuid": source_uuid or "（旧版无项目标识）",
            "units": len(units), "issues": len(issues), "attachments": int(attachment_count),
        }
        sources.append(source)
        fingerprints.append({
            "path": str(path.resolve()), "size": stat.st_size, "mtime_ns": stat.st_mtime_ns,
            "project_uuid": source_uuid, "units": len(units), "issues": len(issues), "attachments": int(attachment_count),
        })

    fingerprint = hashlib.sha256(json.dumps(
        fingerprints, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")).hexdigest()
    return {
        "ok": not blockers, "blockers": blockers, "conflicts": conflicts, "sources": sources,
        "fingerprint": fingerprint, "target_fingerprint": target_fingerprint,
    }


def _import_folder_dir(proj, unit_id, src_dir, folder_name, operator):
    """把备份中的文件夹实体（目录）复制进当前项目附件库并登记记录。"""
    import uuid as _uuid

    from database import ATTACH_DIR
    unit = proj.get_unit(unit_id)
    if not unit:
        return None
    dirname = f"{_safe(folder_name)}_{_uuid.uuid4().hex[:8]}"
    dest = proj.root / ATTACH_DIR / proj.unit_dir_name(unit_id) / dirname
    try:
        shutil.copytree(src_dir, dest)
        total = sum(p.stat().st_size for p in dest.rglob("*") if p.is_file())
        rel = f"{ATTACH_DIR}/{proj.unit_dir_name(unit_id)}/{dirname}"
        conn = proj._conn
        with proj._lock, conn:
            cur = conn.execute(
                "INSERT INTO files(unit_id, stored_name, orig_name, rel_path, size, sha256, mime, created_at) "
                "VALUES(?,?,?,?,?,?,?,?)",
                (unit_id, dirname, folder_name, rel, total, "", "folder",
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            fid = cur.lastrowid
    except Exception:
        shutil.rmtree(dest, ignore_errors=True)
        raise
    proj.log(operator, "导入文件夹", f"{unit['name']} · {folder_name}", "合并导入")
    return proj.get_file(fid)
